"""
app.py v5 — Aplikasi Susut Energi API
Flask berjalan sebagai REST API backend untuk frontend terpisah.
"""

from flask import Flask, Response, g, jsonify, request, send_file
from flask_migrate import Migrate
from .config import Config
from .core.auth import (
    current_user as _current_user,
    logout_user as _logout_user,
    validate_csrf as _validate_csrf,
)
from .core.constants import (
    PUBLIC_ENDPOINTS,
    SAFE_METHODS,
    WRITE_ROLES,
)
from .core.security import (
    audit as _audit,
    json_error as _json_error,
    request_payload as _request_payload,
    validate_password_policy as _validate_password_policy,
)
from .models import (db, GarduInduk, Trafo, Penyulang,
                     MeterReading, FeederReading,
                     TransferAntarUnit, RekapBulanan,
                     EximRule, EximMonthlyResult,
                     User, AuditLog, AreaUnit)
from .routes.system import system_bp
from .routes.auth import auth_bp
from .routes.dashboard import dashboard_bp
from .routes.export import export_bp
from .routes.readings import readings_bp
from .routes.master import master_bp
from .routes.profile import profile_bp
from .routes.security import security_bp
from .routes.upload import upload_bp
from .routes.workflow import workflow_bp
from .routes.kwh_jual import kwh_jual_bp
from .routes.penyulang_area import register_penyulang_area_route
from sqlalchemy import func, text, inspect
from sqlalchemy import and_
from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import click
import io
import os

app = Flask(__name__, static_folder=None)
app.config.from_object(Config)
Config.validate()
app.permanent_session_lifetime = timedelta(hours=app.config.get('PERMANENT_SESSION_HOURS', 8))
db.init_app(app)
MIGRATIONS_DIR = Path(__file__).resolve().parent / 'migrations'
migrate = Migrate(
    app,
    db,
    directory=str(MIGRATIONS_DIR),
    compare_type=True,
    render_as_batch=True,
)

app.register_blueprint(system_bp)
app.register_blueprint(auth_bp)
app.register_blueprint(dashboard_bp)
app.register_blueprint(export_bp)
app.register_blueprint(readings_bp)
register_penyulang_area_route(master_bp)
app.register_blueprint(master_bp)
app.register_blueprint(profile_bp)
app.register_blueprint(security_bp)
app.register_blueprint(upload_bp)
app.register_blueprint(workflow_bp)
app.register_blueprint(kwh_jual_bp)

# ════════════════════════════════════════════════
# KONFIGURASI DATA LOKAL
# ════════════════════════════════════════════════

def _next_month(period):
    return date(period.year + 1, 1, 1) if period.month == 12 else date(period.year, period.month + 1, 1)


def _month_date(value, fallback=None):
    if value is None or str(value).strip() == '':
        if fallback:
            year, month = fallback.split('-')[:2]
            return date(int(year), int(month), 1)
        raise ValueError('Kolom bulan/periode wajib diisi')
    if isinstance(value, datetime):
        return date(value.year, value.month, 1)
    if isinstance(value, date):
        return date(value.year, value.month, 1)

    raw = str(value).strip()
    if len(raw) == 7 and raw[4] == '-':
        return date(int(raw[:4]), int(raw[5:7]), 1)
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%Y', '%m-%Y'):
        try:
            parsed = datetime.strptime(raw, fmt)
            return date(parsed.year, parsed.month, 1)
        except ValueError:
            continue
    raise ValueError(f'Format bulan tidak dikenali: {raw}')


def _report_period_bounds(default_month=False):
    bulan = (request.args.get('bulan') or request.args.get('periode') or '').strip()
    tahun = request.args.get('tahun', type=int) or date.today().year
    bulan_int = request.args.get('month', type=int)
    if bulan:
        period = _month_date(bulan)
        return period, _next_month(period), period.strftime('%Y-%m')
    if bulan_int:
        period = date(tahun, bulan_int, 1)
        return period, _next_month(period), period.strftime('%Y-%m')
    if default_month:
        period = date.today().replace(day=1)
        return period, _next_month(period), period.strftime('%Y-%m')
    return date(tahun, 1, 1), date(tahun + 1, 1, 1), str(tahun)


def _month_filter(query, column, start, end):
    return query.filter(column >= start, column < end)


def _kwh_sum(*columns):
    expr = 0
    for column in columns:
        if isinstance(column, type) and issubclass(column, db.Model):
            continue
        expr += func.coalesce(column, 0)
    return expr


def _float_value(value):
    return float(value or 0)


def _report_dataset(module):
    module = module.replace('-', '_').lower()
    if module in {'rekap', 'rekap_kwh'}:
        return _report_rekap_kwh()
    if module in {'deviasi', 'deviasi_gi'}:
        return _report_deviasi_gi()
    if module == 'proporsional':
        return _report_proporsional()
    if module in {'transfer_exim', 'exim'}:
        return _report_transfer_exim()
    if module in {'transfer_uid', 'transfer_antar_uid'}:
        return _report_transfer_uid()
    raise ValueError('Modul export tidak dikenali.')


def _report_rekap_kwh():
    start, end, period_label = _report_period_bounds()
    q = db.session.query(RekapBulanan, GarduInduk, Trafo).join(
        GarduInduk, RekapBulanan.gi_id == GarduInduk.id
    ).outerjoin(Trafo, RekapBulanan.trafo_id == Trafo.id)
    q = _month_filter(q, RekapBulanan.periode_bulan, start, end)
    headers = ['Periode', 'Gardu Induk', 'Trafo', 'MU Total', 'MP Total', 'Penyulang Total', 'Dev MU-MP %', 'Dev MU-Penyulang %', 'Susut kWh', 'Susut %', 'Ekspor', 'Impor']
    rows = []
    for rekap, gi, trafo in q.order_by(RekapBulanan.periode_bulan, GarduInduk.nama_gi, Trafo.kode_trafo).all():
        rows.append([
            rekap.periode_bulan.strftime('%Y-%m'),
            gi.nama_gi,
            trafo.nama_trafo if trafo else 'TOTAL GI',
            _float_value(rekap.kwh_mu_total),
            _float_value(rekap.kwh_mp_total),
            _float_value(rekap.kwh_penyulang_total),
            _float_value(rekap.deviasi_mu_mp),
            _float_value(rekap.deviasi_mu_penyulang),
            _float_value(rekap.susut_kwh),
            _float_value(rekap.susut_persen),
            _float_value(rekap.transfer_ekspor),
            _float_value(rekap.transfer_impor),
        ])
    return 'Rekap kWh', f'Periode {period_label}', headers, rows, f'rekap_kwh_{period_label}'


def _report_deviasi_gi():
    start, end, period_label = _report_period_bounds()
    gi_id = request.args.get('gi_id', type=int)
    meter_expr = _kwh_sum(MeterReading, MeterReading.mu_kwh_wbp, MeterReading.mu_kwh_lwbp1, MeterReading.mu_kwh_lwbp2)
    mp_expr = _kwh_sum(MeterReading, MeterReading.mp_kwh_wbp, MeterReading.mp_kwh_lwbp1, MeterReading.mp_kwh_lwbp2)
    feeder_expr = _kwh_sum(FeederReading, FeederReading.kwh_wbp, FeederReading.kwh_lwbp1, FeederReading.kwh_lwbp2)
    feeder_sub = db.session.query(
        FeederReading.trafo_id,
        FeederReading.periode_bulan,
        func.sum(feeder_expr).label('feeder_total'),
    ).filter(
        FeederReading.periode_bulan >= start,
        FeederReading.periode_bulan < end,
    ).group_by(FeederReading.trafo_id, FeederReading.periode_bulan).subquery()
    q = db.session.query(
        MeterReading.periode_bulan,
        GarduInduk.nama_gi,
        Trafo.kode_trafo,
        Trafo.nama_trafo,
        meter_expr.label('mu_total'),
        mp_expr.label('mp_total'),
        func.coalesce(feeder_sub.c.feeder_total, 0).label('feeder_total'),
    ).join(Trafo, MeterReading.trafo_id == Trafo.id).join(
        GarduInduk, MeterReading.gi_id == GarduInduk.id
    ).outerjoin(
        feeder_sub,
        (feeder_sub.c.trafo_id == MeterReading.trafo_id) &
        (feeder_sub.c.periode_bulan == MeterReading.periode_bulan),
    )
    q = _month_filter(q, MeterReading.periode_bulan, start, end)
    if gi_id:
        q = q.filter(MeterReading.gi_id == gi_id)
    headers = ['Periode', 'Gardu Induk', 'Kode Trafo', 'Trafo', 'MU Total', 'MP Total', 'Penyulang Total', 'Dev MU-Penyulang kWh', 'Dev MU-Penyulang %', 'Dev MU-MP %']
    rows = []
    for row in q.order_by(MeterReading.periode_bulan, GarduInduk.nama_gi, Trafo.kode_trafo).all():
        mu = _float_value(row.mu_total)
        mp = _float_value(row.mp_total)
        feeder = _float_value(row.feeder_total)
        gap = mu - feeder
        rows.append([
            row.periode_bulan.strftime('%Y-%m'),
            row.nama_gi,
            row.kode_trafo,
            row.nama_trafo,
            mu,
            mp,
            feeder,
            gap,
            (gap / mu * 100) if mu else 0,
            ((mu - mp) / mu * 100) if mu else 0,
        ])
    return 'Deviasi GI', f'Periode {period_label}', headers, rows, f'deviasi_gi_{period_label}'


def _report_proporsional():
    start, end, period_label = _report_period_bounds(default_month=True)
    gi_id = request.args.get('gi_id', type=int)
    mu_expr = _kwh_sum(MeterReading, MeterReading.mu_kwh_wbp, MeterReading.mu_kwh_lwbp1, MeterReading.mu_kwh_lwbp2)
    feeder_expr = _kwh_sum(FeederReading, FeederReading.kwh_wbp, FeederReading.kwh_lwbp1, FeederReading.kwh_lwbp2)
    mu_rows = db.session.query(MeterReading.gi_id, func.sum(mu_expr).label('mu_total')).filter(
        MeterReading.periode_bulan >= start,
        MeterReading.periode_bulan < end,
    )
    if gi_id:
        mu_rows = mu_rows.filter(MeterReading.gi_id == gi_id)
    mu_by_gi = {row.gi_id: _float_value(row.mu_total) for row in mu_rows.group_by(MeterReading.gi_id).all()}
    q = db.session.query(FeederReading, Penyulang, Trafo, GarduInduk).join(
        Penyulang, FeederReading.penyulang_id == Penyulang.id
    ).join(Trafo, FeederReading.trafo_id == Trafo.id).join(
        GarduInduk, FeederReading.gi_id == GarduInduk.id
    )
    q = _month_filter(q, FeederReading.periode_bulan, start, end)
    if gi_id:
        q = q.filter(FeederReading.gi_id == gi_id)
    raw_rows = q.order_by(GarduInduk.nama_gi, Trafo.kode_trafo, Penyulang.kode_penyulang).all()
    feeder_total_by_gi = defaultdict(float)
    for reading, penyulang, trafo, gi in raw_rows:
        feeder_total_by_gi[gi.id] += reading.kwh_total
    headers = ['Periode', 'Gardu Induk', 'Trafo', 'Area/UP3', 'Penyulang', 'Hasil Baca', 'Porsi %', 'Deviasi Dibagi', 'Total Proporsional']
    rows = []
    for reading, penyulang, trafo, gi in raw_rows:
        total_feeder_gi = feeder_total_by_gi.get(gi.id, 0)
        mu_total = mu_by_gi.get(gi.id, 0)
        deviasi = mu_total - total_feeder_gi
        porsi = (reading.kwh_total / total_feeder_gi) if total_feeder_gi else 0
        alokasi = deviasi * porsi
        rows.append([
            reading.periode_bulan.strftime('%Y-%m'),
            gi.nama_gi,
            trafo.kode_trafo,
            penyulang.area_up3 or 'Belum Dipetakan',
            penyulang.nama_penyulang,
            reading.kwh_total,
            porsi * 100,
            alokasi,
            reading.kwh_total + alokasi,
        ])
    return 'Proporsional', f'Periode {period_label}', headers, rows, f'proporsional_{period_label}'


def _report_transfer_exim():
    start, end, period_label = _report_period_bounds()
    q = db.session.query(EximMonthlyResult, EximRule).join(EximRule, EximMonthlyResult.rule_id == EximRule.id)
    q = _month_filter(q, EximMonthlyResult.periode_bulan, start, end)
    headers = ['Periode', 'Rule', 'Metode', 'UP3 Asal', 'UP3 Tujuan', 'Fungsi', 'Arah', 'Basis kWh', 'WBP', 'LWBP1', 'LWBP2', 'Transfer kWh', 'Porsi %', 'Catatan']
    rows = []
    for result, rule in q.order_by(EximMonthlyResult.periode_bulan, EximMonthlyResult.up3_tujuan).all():
        rows.append([
            result.periode_bulan.strftime('%Y-%m'),
            rule.nama_rule or rule.kode_rule,
            result.metode,
            result.up3_asal,
            result.up3_tujuan,
            result.fungsi,
            result.arah,
            _float_value(result.kwh_basis),
            _float_value(result.kwh_wbp),
            _float_value(result.kwh_lwbp1),
            _float_value(result.kwh_lwbp2),
            _float_value(result.kwh_transfer),
            _float_value(result.porsi) * 100 if result.porsi is not None else 0,
            result.catatan,
        ])
    return 'Transfer EXIM', f'Periode {period_label}', headers, rows, f'transfer_exim_{period_label}'


def _report_transfer_uid():
    start, end, period_label = _report_period_bounds()
    q = _month_filter(TransferAntarUnit.query, TransferAntarUnit.periode_bulan, start, end)
    headers = ['Periode', 'Unit Asal', 'Unit Tujuan', 'GI/Interkoneksi', 'Kode Interbus', 'Arah', 'Transfer kWh', 'Catatan']
    rows = []
    for item in q.order_by(TransferAntarUnit.periode_bulan, TransferAntarUnit.unit_asal).all():
        rows.append([
            item.periode_bulan.strftime('%Y-%m'),
            item.unit_asal,
            item.unit_tujuan,
            item.gi_interkoneksi,
            item.kode_interbus,
            item.arah,
            _float_value(item.kwh_transfer),
            item.catatan,
        ])
    return 'Transfer Antar UID', f'Periode {period_label}', headers, rows, f'transfer_antar_uid_{period_label}'


def _report_filter_label():
    labels = []
    for key, label in [
        ('tahun', 'Tahun'),
        ('month', 'Bulan'),
        ('bulan', 'Periode'),
        ('periode', 'Periode'),
        ('gi_id', 'GI ID'),
    ]:
        value = (request.args.get(key) or '').strip()
        if value:
            labels.append(f'{label}: {value}')
    return ' | '.join(labels) if labels else 'Semua data'


def _sumable_report_column(header):
    text = str(header or '').lower()
    if '%' in text or 'porsi' in text or 'faktor' in text:
        return False
    return any(token in text for token in [
        'kwh', 'wbp', 'lwbp', 'mu', 'mp', 'penyulang', 'susut',
        'transfer', 'ekspor', 'impor', 'hasil baca', 'proporsional',
        'deviasi dibagi', 'basis',
    ])


def _excel_report(title, subtitle, headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Laporan'
    ws.append([title])
    ws.append([subtitle])
    ws.append([f'Dibuat: {datetime.now().strftime("%d/%m/%Y %H:%M")} | Filter: {_report_filter_label()}'])
    ws.append([])
    ws.append(headers)
    for row in rows:
        ws.append(row)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(headers), 1))
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max(len(headers), 1))
    ws['A1'].font = Font(size=14, bold=True)
    ws['A2'].font = Font(size=10, italic=True, color='667085')
    ws['A3'].font = Font(size=9, color='667085')
    header_row = 5
    fill = PatternFill('solid', fgColor='E8F1FF')
    total_fill = PatternFill('solid', fgColor='F3F6FB')
    thin = Side(style='thin', color='D7DEE8')
    border = Border(top=thin, right=thin, bottom=thin, left=thin)
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color='1D2430')
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    first_data_row = header_row + 1
    last_data_row = ws.max_row
    for row_idx in range(first_data_row, last_data_row + 1):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=not isinstance(cell.value, (int, float)))
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00' if ('%' in str(headers[col_idx - 1]) or abs(cell.value) % 1) else '#,##0'

    if rows:
        total_row = ws.max_row + 1
        ws.cell(total_row, 1, 'TOTAL')
        for col_idx, header in enumerate(headers, start=1):
            values = [
                row[col_idx - 1]
                for row in rows
                if len(row) >= col_idx and isinstance(row[col_idx - 1], (int, float))
            ]
            if values and _sumable_report_column(header):
                ws.cell(total_row, col_idx, sum(values))
        for cell in ws[total_row]:
            cell.font = Font(bold=True, color='1D2430')
            cell.fill = total_fill
            cell.border = border
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00' if abs(cell.value) % 1 else '#,##0'

    ws.freeze_panes = 'A6'
    ws.auto_filter.ref = f'A{header_row}:{get_column_letter(max(len(headers), 1))}{max(ws.max_row, header_row)}'
    ws.print_title_rows = f'{header_row}:{header_row}'
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    for col_idx in range(1, len(headers) + 1):
        width = max(
            len(str(ws.cell(row=row_idx, column=col_idx).value or ''))
            for row_idx in range(1, min(ws.max_row, 80) + 1)
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 12), 34)
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def _pdf_escape(value):
    return str(value if value is not None else '').replace('\\', '\\\\').replace('(', '\\(').replace(')', '\\)')


def _pdf_report(title, subtitle, headers, rows):
    text_rows = [' | '.join(headers)]
    text_rows.extend(' | '.join(str(value if value is not None else '') for value in row) for row in rows[:120])
    lines = [
        title,
        subtitle,
        f'Dibuat: {datetime.now().strftime("%d/%m/%Y %H:%M")} | Filter: {_report_filter_label()}',
        f'Total baris: {len(rows)}',
        '',
    ] + [line[:145] for line in text_rows]
    pages = [lines[i:i + 42] for i in range(0, len(lines), 42)] or [[]]
    objects = [
        b'<< /Type /Catalog /Pages 2 0 R >>',
        None,
        b'<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>',
    ]
    page_ids = []
    for page_lines in pages:
        content_lines = ['BT', '/F1 9 Tf']
        y = 800
        for line in page_lines:
            content_lines.append(f'1 0 0 1 36 {y} Tm ({_pdf_escape(line)}) Tj')
            y -= 17
        content_lines.append('ET')
        stream = '\n'.join(content_lines).encode('latin-1', 'replace')
        page_id = len(objects) + 1
        content_id = len(objects) + 2
        page_ids.append(page_id)
        objects.append(f'<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Resources << /Font << /F1 3 0 R >> >> /Contents {content_id} 0 R >>'.encode('latin-1'))
        objects.append(b'<< /Length ' + str(len(stream)).encode('ascii') + b' >>\nstream\n' + stream + b'\nendstream')
    kids = ' '.join(f'{page_id} 0 R' for page_id in page_ids)
    objects[1] = f'<< /Type /Pages /Kids [{kids}] /Count {len(page_ids)} >>'.encode('latin-1')
    output = io.BytesIO()
    output.write(b'%PDF-1.4\n')
    offsets = [0]
    for idx, obj in enumerate(objects, start=1):
        offsets.append(output.tell())
        output.write(f'{idx} 0 obj\n'.encode('ascii'))
        output.write(obj)
        output.write(b'\nendobj\n')
    xref_pos = output.tell()
    output.write(f'xref\n0 {len(objects) + 1}\n'.encode('ascii'))
    output.write(b'0000000000 65535 f \n')
    for offset in offsets[1:]:
        output.write(f'{offset:010d} 00000 n \n'.encode('ascii'))
    output.write(f'trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_pos}\n%%EOF'.encode('ascii'))
    return output.getvalue()


def _report_file_response(module, fmt):
    title, subtitle, headers, rows, filename = _report_dataset(module)
    fmt = fmt.lower()
    if fmt not in {'xlsx', 'pdf'}:
        raise ValueError('Format export harus xlsx atau pdf.')
    _audit('EXPORT_REPORT', entity_type='report', entity_id=module, detail={
        'module': module,
        'format': fmt,
        'rows': len(rows),
        'subtitle': subtitle,
    })
    db.session.commit()
    if fmt == 'xlsx':
        stream = _excel_report(title, subtitle, headers, rows)
        return send_file(
            stream,
            as_attachment=True,
            download_name=f'{filename}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    if fmt == 'pdf':
        pdf_bytes = _pdf_report(title, subtitle, headers, rows)
        return Response(
            pdf_bytes,
            mimetype='application/pdf',
            headers={'Content-Disposition': f'attachment; filename="{filename}.pdf"'},
        )


@app.before_request
def apply_security_gate():
    g.current_user = _current_user()
    if g.current_user and not g.current_user.aktif:
        _logout_user()
        g.current_user = None

    if request.endpoint in PUBLIC_ENDPOINTS:
        return None

    if request.method not in SAFE_METHODS and not _validate_csrf():
        return _json_error('CSRF token tidak valid.', 403)

    if app.config.get('SECURITY_REQUIRE_LOGIN', True) and not g.current_user:
        return _json_error('Login diperlukan.', 401)
    return None


@app.after_request
def add_security_headers(response):
    response.headers.setdefault('X-Content-Type-Options', 'nosniff')
    response.headers.setdefault('X-Frame-Options', 'SAMEORIGIN')
    response.headers.setdefault('Referrer-Policy', 'same-origin')
    response.headers.setdefault('Permissions-Policy', 'camera=(), microphone=(), geolocation=()')
    response.headers.setdefault(
        'Content-Security-Policy',
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' https://cdnjs.cloudflare.com; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com https://cdn.jsdelivr.net; "
        "font-src 'self' https://fonts.gstatic.com https://cdn.jsdelivr.net; "
        "img-src 'self' data:; connect-src 'self'; frame-ancestors 'self';"
    )
    return response


@app.cli.command('create-admin')
@click.option('--username', prompt=True)
@click.option('--password', prompt=True, hide_input=True, confirmation_prompt=True)
@click.option('--name', default='Administrator')
def create_admin_command(username, password, name):
    password_error = _validate_password_policy(password)
    if password_error:
        raise click.ClickException(password_error)
    existing = User.query.filter_by(username=username.strip()).first()
    if existing:
        existing.role = 'admin'
        existing.aktif = True
        existing.nama_lengkap = name or existing.nama_lengkap
        existing.set_password(password)
        action = 'diperbarui'
    else:
        user = User(username=username.strip(), nama_lengkap=name, role='admin', aktif=True)
        user.set_password(password)
        db.session.add(user)
        action = 'dibuat'
    db.session.commit()
    click.echo(f'Admin {username} berhasil {action}.')


# ════════════════════════════════════════════════
# API — MASTER DATA
# ════════════════════════════════════════════════

def _master_writer_required():
    user = getattr(g, 'current_user', None)
    if not user or user.role not in WRITE_ROLES:
        return _json_error('Akses ubah master data hanya untuk admin/operator.', 403)
    return None


def _clean_value(value, default=''):
    text_value = str(value or '').strip()
    return text_value if text_value else default


def _decimal_payload(value, default='0'):
    if value in (None, ''):
        return Decimal(default)
    return Decimal(str(value))


# ════════════════════════════════════════════════
# API — DASHBOARD
# ════════════════════════════════════════════════

# ════════════════════════════════════════════════
# API — FEEDER, METER, TRANSFER, REKAP
# ════════════════════════════════════════════════

@app.route('/api/transfer-data')
def api_transfer_data():
    try:
        bulan = request.args.get('bulan')
        tahun = request.args.get('tahun', type=int)
        q = TransferAntarUnit.query
        if bulan:
            thn, bln = bulan.split('-')
            q = q.filter(
                func.extract('year',  TransferAntarUnit.periode_bulan) == int(thn),
                func.extract('month', TransferAntarUnit.periode_bulan) == int(bln)
            )
        elif tahun:
            q = q.filter(func.extract('year', TransferAntarUnit.periode_bulan) == tahun)
        return jsonify([r.to_dict() for r in q.order_by(TransferAntarUnit.periode_bulan).all()])
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/rekap')
def api_rekap():
    try:
        tahun = request.args.get('tahun', type=int)
        gi_id = request.args.get('gi_id', type=int)
        q = db.session.query(RekapBulanan, GarduInduk)\
              .join(GarduInduk, RekapBulanan.gi_id == GarduInduk.id)\
              .filter(RekapBulanan.trafo_id.is_(None))
        if tahun: q = q.filter(func.extract('year', RekapBulanan.periode_bulan) == tahun)
        if gi_id: q = q.filter(RekapBulanan.gi_id == gi_id)
        result = []
        for rb, gi in q.order_by(RekapBulanan.periode_bulan).all():
            d = rb.to_dict()
            d['nama_gi'] = gi.nama_gi
            d['kode_gi'] = gi.kode_gi
            result.append(d)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ════════════════════════════════════════════════
# ════════════════════════════════════════════════

# ════════════════════════════════════════════════
# ERROR HANDLERS
# ════════════════════════════════════════════════

@app.errorhandler(403)
def forbidden(e):
    return jsonify({
        'error': 'Forbidden',
        'message': str(e),
        'status_code': 403,
    }), 403


@app.errorhandler(413)
def payload_too_large(e):
    max_mb = app.config['MAX_CONTENT_LENGTH'] // (1024 * 1024)
    return jsonify({
        'error': 'Payload Too Large',
        'message': f'Ukuran file melebihi batas {max_mb} MB.',
        'status_code': 413,
    }), 413


@app.errorhandler(404)
def not_found(e):
    return jsonify({
        'error': 'Not Found',
        'message': str(e),
        'status_code': 404,
    }), 404

@app.errorhandler(500)
def server_error(e):
    return jsonify({
        'error': 'Internal Server Error',
        'status_code': 500,
    }), 500


# ════════════════════════════════════════════════
# JALANKAN
# ════════════════════════════════════════════════

if __name__ == '__main__':
    debug_mode = os.getenv('FLASK_DEBUG', 'false').lower() == 'true'
    app.run(debug=debug_mode, host='0.0.0.0', port=5000)

import logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
