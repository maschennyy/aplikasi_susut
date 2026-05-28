from collections import Counter, defaultdict
from datetime import date, datetime
import re

from openpyxl import load_workbook


MONTHS_ID = {
    'JANUARI': 1,
    'JAN': 1,
    'FEBRUARI': 2,
    'FEB': 2,
    'MARET': 3,
    'MAR': 3,
    'APRIL': 4,
    'APR': 4,
    'MEI': 5,
    'JUNI': 6,
    'JUN': 6,
    'JULI': 7,
    'JUL': 7,
    'AGUSTUS': 8,
    'AGS': 8,
    'AUG': 8,
    'SEPTEMBER': 9,
    'SEP': 9,
    'OKTOBER': 10,
    'OKT': 10,
    'NOVEMBER': 11,
    'NOV': 11,
    'DESEMBER': 12,
    'DES': 12,
    'DEC': 12,
}


REGISTER_SUFFIXES = {
    'H': 'wbp',
    'WBP': 'wbp',
    'L1': 'lwbp1',
    'LWBP1': 'lwbp1',
    'LWBP': 'lwbp1',
    'L2': 'lwbp2',
    'LWBP2': 'lwbp2',
}


def clean_text(value, default=''):
    if value is None:
        return default
    text = str(value).replace('\xa0', ' ').strip()
    return re.sub(r'\s+', ' ', text) if text else default


def numeric(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = clean_text(value)
    if not text or text in {'-', '--'}:
        return None
    text = text.replace('.', '').replace(',', '.')
    try:
        return float(text)
    except ValueError:
        return None


def parse_period(value):
    if isinstance(value, datetime):
        return date(value.year, value.month, 1)
    if isinstance(value, date):
        return date(value.year, value.month, 1)
    text = clean_text(value).upper()
    if not text:
        return None
    year_match = re.search(r'(20\d{2}|19\d{2})', text)
    month = None
    for label, number in MONTHS_ID.items():
        if re.search(rf'\b{re.escape(label)}\b', text):
            month = number
            break
    if year_match and month:
        return date(int(year_match.group(1)), month, 1)
    iso_match = re.search(r'(20\d{2})[-/](\d{1,2})', text)
    if iso_match:
        return date(int(iso_match.group(1)), int(iso_match.group(2)), 1)
    return None


def open_workbook(source, data_only=True):
    if hasattr(source, 'sheetnames'):
        return source
    if hasattr(source, 'seek'):
        source.seek(0)
    return load_workbook(source, data_only=data_only, read_only=False)


def register_from_name(name):
    parts = clean_text(name).upper().split()
    if not parts:
        return None
    return REGISTER_SUFFIXES.get(parts[-1])


def feeder_base_name(name):
    text = clean_text(name)
    parts = text.split()
    if parts and parts[-1].upper() in REGISTER_SUFFIXES:
        return ' '.join(parts[:-1]).strip()
    return text


def feeder_code(name):
    text = feeder_base_name(name).upper()
    text = re.sub(r'[^A-Z0-9]+', '-', text).strip('-')
    return text[:30] or 'PENYULANG'


def parse_nkwh_feeders(source):
    wb = open_workbook(source, data_only=True)
    if 'kWh Penyulang' not in wb.sheetnames:
        raise ValueError('Sheet "kWh Penyulang" tidak ditemukan.')

    ws = wb['kWh Penyulang']
    period = parse_period(ws['H1'].value) or parse_period(ws['G1'].value)
    raw_rows = []
    current_gi = ''

    for row_idx in range(4, ws.max_row + 1):
        gi_cell = clean_text(ws.cell(row_idx, 1).value)
        feeder_name = clean_text(ws.cell(row_idx, 2).value)
        register = register_from_name(feeder_name)

        if gi_cell and gi_cell.upper() not in {'GI', 'GARDU INDUK', 'NO'}:
            current_gi = gi_cell

        if not feeder_name or not register:
            continue

        base = feeder_base_name(feeder_name)
        if not base or base.upper() in {'NAMA PENYULANG', 'PENYULANG'}:
            continue

        raw_rows.append({
            'row': row_idx,
            'gi': gi_cell if gi_cell.upper() not in {'GI', 'GARDU INDUK', 'NO'} else '',
            'current_gi': current_gi,
            'feeder_name': base,
            'register': register,
            'trafo': clean_text(ws.cell(row_idx, 3).value),
            'faktor_kali': numeric(ws.cell(row_idx, 4).value),
            'stand_awal': numeric(ws.cell(row_idx, 5).value),
            'stand_akhir': numeric(ws.cell(row_idx, 6).value),
            'kwh_meter': numeric(ws.cell(row_idx, 7).value),
            'kwh_manual': numeric(ws.cell(row_idx, 8).value),
        })

    groups = []
    pending = None
    last_gi = ''

    def flush_pending():
        nonlocal pending, last_gi
        if not pending:
            return
        gi_name = pending['gi'] or last_gi or 'Belum Dipetakan'
        if pending['gi']:
            last_gi = pending['gi']
        registers = pending['registers']
        kwh = {}
        manual = {}
        stands = {}
        for reg, row in registers.items():
            manual_value = row.get('kwh_manual')
            meter_value = row.get('kwh_meter')
            kwh[reg] = manual_value if manual_value is not None else (meter_value or 0)
            manual[reg] = manual_value
            stands[reg] = {
                'stand_awal': row.get('stand_awal'),
                'stand_akhir': row.get('stand_akhir'),
                'faktor_kali': row.get('faktor_kali') or 1,
                'kwh_meter': meter_value,
            }
        groups.append({
            'kode_penyulang': feeder_code(pending['feeder_name']),
            'nama_penyulang': pending['feeder_name'],
            'gardu_induk': gi_name,
            'kode_trafo': pending['trafo'] or 'TRF-1',
            'nama_trafo': f"Trafo {pending['trafo']}" if pending['trafo'] else 'Trafo 1',
            'periode_bulan': period.isoformat() if period else None,
            'kwh_wbp': kwh.get('wbp', 0),
            'kwh_lwbp1': kwh.get('lwbp1', 0),
            'kwh_lwbp2': kwh.get('lwbp2', 0),
            'kwh_total': sum(kwh.values()),
            'manual_kwh_wbp': manual.get('wbp'),
            'manual_kwh_lwbp1': manual.get('lwbp1'),
            'manual_kwh_lwbp2': manual.get('lwbp2'),
            'registers': stands,
            'source_sheet': ws.title,
            'source_row_start': pending['row_start'],
            'source_row_end': pending['row_end'],
        })
        pending = None

    for row in raw_rows:
        if not pending or pending['feeder_name'] != row['feeder_name']:
            flush_pending()
            pending = {
                'feeder_name': row['feeder_name'],
                'gi': row['gi'] or row['current_gi'],
                'trafo': row['trafo'],
                'registers': {},
                'row_start': row['row'],
                'row_end': row['row'],
            }
        if row['gi']:
            pending['gi'] = row['gi']
        if row['trafo']:
            pending['trafo'] = row['trafo']
        pending['registers'][row['register']] = row
        pending['row_end'] = row['row']
    flush_pending()

    gi_counter = Counter(item['gardu_induk'] for item in groups)
    trafo_counter = Counter((item['gardu_induk'], item['kode_trafo']) for item in groups)
    total_kwh = sum(item['kwh_total'] for item in groups)

    return {
        'periode_bulan': period.isoformat() if period else None,
        'sheet': ws.title,
        'register_rows': len(raw_rows),
        'feeder_count': len(groups),
        'gi_count': len(gi_counter),
        'trafo_count': len(trafo_counter),
        'total_kwh': total_kwh,
        'by_gi': [{'gardu_induk': key, 'jumlah_penyulang': value}
                  for key, value in gi_counter.most_common()],
        'feeders': groups,
    }


def parse_tng_structure(source):
    wb = open_workbook(source, data_only=True)
    if 'TNG' not in wb.sheetnames:
        return {'gi_blocks': [], 'sections': []}
    ws = wb['TNG']
    gi_blocks = []
    sections = Counter()
    current_gi = ''
    for row_idx in range(1, ws.max_row + 1):
        a = clean_text(ws.cell(row_idx, 1).value)
        b = clean_text(ws.cell(row_idx, 2).value)
        c = clean_text(ws.cell(row_idx, 3).value)
        if a.upper().startswith('GARDU INDUK') and c:
            current_gi = c
            gi_blocks.append({'gardu_induk': c, 'row': row_idx})
        label = a or b
        upper = label.upper()
        if current_gi and re.match(r'^(I|II|III|IV|V|VI|VII|VIII)(\.|\s)', upper):
            sections[upper] += 1
    return {
        'gi_blocks': gi_blocks,
        'sections': [{'label': key, 'count': value} for key, value in sections.items()],
    }


def parse_exim_rows(source):
    wb = open_workbook(source, data_only=True)
    if 'Exim' not in wb.sheetnames:
        return {'sheet': 'Exim', 'row_count': 0, 'methods': [], 'rows': []}
    ws = wb['Exim']
    rows = []
    current_gi = ''
    for row_idx in range(8, ws.max_row + 1):
        gi = clean_text(ws.cell(row_idx, 2).value)
        feeder = clean_text(ws.cell(row_idx, 4).value)
        if gi:
            current_gi = gi
        if not feeder:
            continue

        area_asal = clean_text(ws.cell(row_idx, 9).value)
        fungsi = clean_text(ws.cell(row_idx, 10).value).upper()
        area_tujuan = clean_text(ws.cell(row_idx, 11).value)
        direct_values = [numeric(ws.cell(row_idx, col).value) for col in range(14, 24)]
        manual_values = [numeric(ws.cell(row_idx, col).value) for col in range(24, 28)]
        kva_values = [numeric(ws.cell(row_idx, col).value) for col in range(30, 35)]
        total_direct = numeric(ws.cell(row_idx, 23).value)
        total_manual = numeric(ws.cell(row_idx, 27).value)

        if any(value is not None for value in kva_values):
            method = 'KVA_PROPORSI'
        elif any(value is not None for value in direct_values):
            method = 'DIRECT_STAND'
        elif total_manual is not None or any(value is not None for value in manual_values):
            method = 'KWH_JUAL'
        else:
            method = 'ADJUSTMENT'

        rows.append({
            'row': row_idx,
            'gardu_induk': gi or current_gi or 'Belum Dipetakan',
            'feeder': feeder,
            'lokasi': clean_text(ws.cell(row_idx, 6).value),
            'jenis': clean_text(ws.cell(row_idx, 8).value),
            'area_asal': area_asal,
            'fungsi': fungsi,
            'area_tujuan': area_tujuan,
            'metode': method,
            'arah': 'EKSPOR' if fungsi == 'KIRIM' else 'IMPOR',
            'faktor_meter': numeric(ws.cell(row_idx, 12).value),
            'faktor_kali': numeric(ws.cell(row_idx, 13).value),
            'kwh_wbp': numeric(ws.cell(row_idx, 24).value) or numeric(ws.cell(row_idx, 20).value) or 0,
            'kwh_lwbp1': numeric(ws.cell(row_idx, 25).value) or numeric(ws.cell(row_idx, 21).value) or 0,
            'kwh_lwbp2': numeric(ws.cell(row_idx, 26).value) or numeric(ws.cell(row_idx, 22).value) or 0,
            'kwh_total': total_manual if total_manual is not None else (total_direct or 0),
            'kva_pemilik': numeric(ws.cell(row_idx, 30).value),
            'kva_penerima_1': numeric(ws.cell(row_idx, 31).value),
            'kva_penerima_2': numeric(ws.cell(row_idx, 32).value),
            'kva_total': numeric(ws.cell(row_idx, 33).value),
            'kwh_penyulang_basis': numeric(ws.cell(row_idx, 34).value),
        })

    methods = Counter(item['metode'] for item in rows)
    by_area = defaultdict(float)
    for item in rows:
        key = item['area_tujuan'] or item['area_asal'] or 'Belum Dipetakan'
        by_area[key] += item['kwh_total'] or 0

    return {
        'sheet': ws.title,
        'row_count': len(rows),
        'methods': [{'metode': key, 'count': value} for key, value in methods.items()],
        'by_area': [{'area': key, 'kwh_total': value} for key, value in sorted(by_area.items())],
        'rows': rows,
    }


def analyze_workbook(source):
    wb = open_workbook(source, data_only=True)
    sheet_summary = []
    for ws in wb.worksheets:
        used_rows = ws.max_row
        used_cols = ws.max_column
        for row_idx in range(1, min(ws.max_row, 60) + 1):
            for col_idx in range(1, min(ws.max_column, 80) + 1):
                if ws.cell(row_idx, col_idx).value is not None:
                    used_rows = max(used_rows, row_idx)
                    used_cols = max(used_cols, col_idx)
        sheet_summary.append({
            'name': ws.title,
            'max_row': ws.max_row,
            'max_column': ws.max_column,
            'used_rows': used_rows,
            'used_columns': used_cols,
        })

    feeders = parse_nkwh_feeders(wb) if 'kWh Penyulang' in wb.sheetnames else None
    tng = parse_tng_structure(wb)
    exim = parse_exim_rows(wb)

    return {
        'workbook': {
            'sheet_count': len(wb.sheetnames),
            'sheets': sheet_summary,
        },
        'periode_bulan': feeders.get('periode_bulan') if feeders else None,
        'kwh_penyulang': {
            key: value for key, value in (feeders or {}).items()
            if key != 'feeders'
        },
        'tng': tng,
        'exim': {
            key: value for key, value in exim.items()
            if key != 'rows'
        },
        'samples': {
            'feeders': (feeders or {}).get('feeders', [])[:8],
            'exim': exim.get('rows', [])[:8],
        },
    }
