"""
app.py v5 — Aplikasi Susut Energi
Semua route yang direferensikan di base.html sudah ada.
"""

from flask import (Flask, Response, abort, g, jsonify, redirect,
                   render_template, request, send_file, session, url_for)
from config import Config
from models import (db, GarduInduk, Trafo, Penyulang,
                    MeterReading, FeederReading,
                    TransferAntarUnit, RekapBulanan,
                    EximRule, EximMonthlyResult,
                    User, AuditLog, AreaUnit, MonthlyDataStatus,
                    KwhJual)
from nkwh_excel import analyze_workbook, parse_nkwh_feeders, parse_exim_rows
from sqlalchemy import func, text, inspect
from sqlalchemy import and_
from collections import defaultdict, deque
from datetime import date, datetime, timedelta
from decimal import Decimal
from functools import wraps
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import click
import io
import json
import pandas as pd
import secrets
import os

app = Flask(__name__)
app.config.from_object(Config)
Config.validate()
app.permanent_session_lifetime = timedelta(hours=app.config.get('PERMANENT_SESSION_HOURS', 8))
db.init_app(app)

with app.app_context():
    db.create_all()
    inspector = inspect(db.engine)
    table_columns = {
        table: {col['name'] for col in inspector.get_columns(table)}
        for table in inspector.get_table_names()
    }
    schema_additions = {
        'penyulang': [
            ('area_up3', 'VARCHAR(100)'),
            ('ex_cabang', 'VARCHAR(50)'),
            ('status', "VARCHAR(30) DEFAULT 'AKTIF'"),
        ],
        'feeder_reading': [
            ('deviasi_persen', 'NUMERIC(8, 2)'),
            ('anomaly_type', 'VARCHAR(30)'),
            ('wbp_stand_awal', 'NUMERIC(15, 2)'),
            ('wbp_stand_akhir', 'NUMERIC(15, 2)'),
            ('wbp_faktor_kali', 'NUMERIC(10, 2)'),
            ('lwbp1_stand_awal', 'NUMERIC(15, 2)'),
            ('lwbp1_stand_akhir', 'NUMERIC(15, 2)'),
            ('lwbp1_faktor_kali', 'NUMERIC(10, 2)'),
            ('lwbp2_stand_awal', 'NUMERIC(15, 2)'),
            ('lwbp2_stand_akhir', 'NUMERIC(15, 2)'),
            ('lwbp2_faktor_kali', 'NUMERIC(10, 2)'),
            ('manual_kwh_wbp', 'NUMERIC(15, 2)'),
            ('manual_kwh_lwbp1', 'NUMERIC(15, 2)'),
            ('manual_kwh_lwbp2', 'NUMERIC(15, 2)'),
            ('source_format', 'VARCHAR(30)'),
            ('source_sheet', 'VARCHAR(80)'),
            ('source_row_start', 'INTEGER'),
            ('source_row_end', 'INTEGER'),
        ],
    }
    for table, columns in schema_additions.items():
        existing = table_columns.get(table, set())
        for column_name, ddl in columns:
            if column_name not in existing:
                db.session.execute(text(f'ALTER TABLE {table} ADD COLUMN {column_name} {ddl}'))
    if User.query.count() == 0:
        admin_username = os.getenv('ADMIN_USERNAME')
        admin_password = os.getenv('ADMIN_PASSWORD')
        if admin_username and admin_password:
            admin = User(
                username=admin_username.strip(),
                nama_lengkap=os.getenv('ADMIN_NAME', 'Administrator'),
                email=os.getenv('ADMIN_EMAIL'),
                role='admin',
                aktif=True,
            )
            admin.set_password(admin_password)
            db.session.add(admin)
    db.session.commit()


# ════════════════════════════════════════════════
# HALAMAN — semua route yang ada di sidebar
# ════════════════════════════════════════════════

SAFE_METHODS = {'GET', 'HEAD', 'OPTIONS'}
PUBLIC_ENDPOINTS = {'login', 'static'}
WRITE_ROLES = {'admin', 'operator'}
ALLOWED_GENERIC_UPLOADS = {'csv', 'xlsx', 'xlsm', 'xls'}
ALLOWED_NKWH_UPLOADS = {'xlsx', 'xlsm'}
XLS_SIGNATURE = bytes.fromhex('d0cf11e0a1b11ae1')
LOGIN_FAILURES = defaultdict(deque)
LOGIN_LOCKOUTS = {}
UPLOAD_EVENTS = defaultdict(deque)
ROLES = {'admin', 'operator', 'viewer', 'auditor'}
MODULE_ACCESS_MATRIX = [
    {
        'module': 'Dashboard',
        'group': 'Dashboard',
        'access': {
            'read': ['viewer', 'auditor', 'operator', 'admin'],
            'export': ['auditor', 'operator', 'admin'],
        },
    },
    {
        'module': 'Gardu Induk',
        'group': 'Gardu Induk',
        'access': {
            'read': ['viewer', 'auditor', 'operator', 'admin'],
            'write': ['operator', 'admin'],
            'export': ['auditor', 'operator', 'admin'],
            'finalize': ['operator', 'admin'],
            'lock': ['admin'],
        },
    },
    {
        'module': 'UID',
        'group': 'UID',
        'access': {
            'read': ['viewer', 'auditor', 'operator', 'admin'],
            'write': ['operator', 'admin'],
            'export': ['auditor', 'operator', 'admin'],
        },
    },
    {
        'module': 'Master Data',
        'group': 'Master',
        'access': {
            'read': ['operator', 'admin'],
            'write': ['operator', 'admin'],
            'audit': ['admin'],
        },
    },
    {
        'module': 'Rekap kWh',
        'group': 'Master',
        'access': {
            'read': ['viewer', 'auditor', 'operator', 'admin'],
            'export': ['auditor', 'operator', 'admin'],
        },
    },
    {
        'module': 'Transaksi',
        'group': 'Transaksi',
        'access': {
            'read': ['viewer', 'auditor', 'operator', 'admin'],
            'write': ['operator', 'admin'],
        },
    },
    {
        'module': 'Security',
        'group': 'Admin',
        'access': {
            'read': ['admin'],
            'write': ['admin'],
            'audit': ['admin'],
        },
    },
    {
        'module': 'Profile',
        'group': 'Akun',
        'access': {
            'read': ['viewer', 'auditor', 'operator', 'admin'],
            'self_update': ['viewer', 'auditor', 'operator', 'admin'],
        },
    },
]
WORKFLOW_STATUS_ORDER = ['DRAFT', 'SUDAH_UPLOAD', 'SUDAH_DICEK', 'FINAL', 'TERKUNCI']
WORKFLOW_STATUS_LABELS = {
    'DRAFT': 'Draft',
    'SUDAH_UPLOAD': 'Sudah Upload',
    'SUDAH_DICEK': 'Sudah Dicek',
    'FINAL': 'Final',
    'TERKUNCI': 'Terkunci',
}
WORKFLOW_TRANSITIONS = {
    'DRAFT': ['DRAFT', 'SUDAH_UPLOAD'],
    'SUDAH_UPLOAD': ['DRAFT', 'SUDAH_UPLOAD', 'SUDAH_DICEK'],
    'SUDAH_DICEK': ['SUDAH_UPLOAD', 'SUDAH_DICEK', 'FINAL'],
    'FINAL': ['SUDAH_DICEK', 'FINAL', 'TERKUNCI'],
    'TERKUNCI': ['TERKUNCI', 'FINAL'],
}
WORKFLOW_WRITABLE_STATUSES = {'DRAFT', 'SUDAH_UPLOAD'}
MONTHLY_ACTIVITY_ACTIONS = [
    'ANALYZE_NKWH',
    'IMPORT_NKWH',
    'IMPORT_PENYULANG',
    'MARK_MONTH_UPLOADED',
    'UPDATE_MONTHLY_STATUS',
]
KWH_JUAL_GROUP_LABELS = {
    'S': 'Sosial',
    'R': 'Rumah Tangga',
    'B': 'Bisnis',
    'I': 'Industri',
    'P': 'Pemerintah',
    'TCL': 'T/C/L Khusus',
}
KWH_JUAL_CATALOG = [
    {'group': 'S', 'golongan': 'S', 'sub_golongan': 'S.1 / 450 VA', 'tegangan': 'TR'},
    {'group': 'S', 'golongan': 'S', 'sub_golongan': 'S.1 / 900 VA', 'tegangan': 'TR'},
    {'group': 'S', 'golongan': 'S', 'sub_golongan': 'S.1 / 1.300 VA', 'tegangan': 'TR'},
    {'group': 'S', 'golongan': 'S', 'sub_golongan': 'S.1 / 2.200 VA', 'tegangan': 'TR'},
    {'group': 'S', 'golongan': 'S', 'sub_golongan': 'S.1 / 3.500 VA s.d 200 kVA', 'tegangan': 'TR'},
    {'group': 'S', 'golongan': 'S', 'sub_golongan': 'S.2 / > 200 kVA s.d < 30.000 kVA', 'tegangan': 'TM'},
    {'group': 'R', 'golongan': 'R', 'sub_golongan': 'R.1 / 450 VA', 'tegangan': 'TR'},
    {'group': 'R', 'golongan': 'R', 'sub_golongan': 'R.1 / 900 VA', 'tegangan': 'TR'},
    {'group': 'R', 'golongan': 'R', 'sub_golongan': 'R.1M / 900 VA', 'tegangan': 'TR'},
    {'group': 'R', 'golongan': 'R', 'sub_golongan': 'R.1 / 1.300 VA', 'tegangan': 'TR'},
    {'group': 'R', 'golongan': 'R', 'sub_golongan': 'R.1 / 2.200 VA', 'tegangan': 'TR'},
    {'group': 'R', 'golongan': 'R', 'sub_golongan': 'R.2 / 3.500 VA s.d 5.500 VA', 'tegangan': 'TR'},
    {'group': 'R', 'golongan': 'R', 'sub_golongan': 'R.3 / 6.600 VA s.d 200 kVA', 'tegangan': 'TR'},
    {'group': 'R', 'golongan': 'R', 'sub_golongan': 'R.3 / > 200 kVA s.d < 30.000 kVA', 'tegangan': 'TM'},
    {'group': 'B', 'golongan': 'B', 'sub_golongan': 'B.1 / 450 VA', 'tegangan': 'TR'},
    {'group': 'B', 'golongan': 'B', 'sub_golongan': 'B.1 / 900 VA', 'tegangan': 'TR'},
    {'group': 'B', 'golongan': 'B', 'sub_golongan': 'B.1 / 1.300 VA', 'tegangan': 'TR'},
    {'group': 'B', 'golongan': 'B', 'sub_golongan': 'B.1 / 2.200 VA s.d 5.500 VA', 'tegangan': 'TR'},
    {'group': 'B', 'golongan': 'B', 'sub_golongan': 'B.2 / 6.600 VA s.d 200 kVA', 'tegangan': 'TR'},
    {'group': 'B', 'golongan': 'B', 'sub_golongan': 'B.3 / > 200 kVA s.d < 30.000 kVA', 'tegangan': 'TM'},
    {'group': 'B', 'golongan': 'B', 'sub_golongan': 'B.3 / 30.000 kVA keatas', 'tegangan': 'TT'},
    {'group': 'I', 'golongan': 'I', 'sub_golongan': 'I.1 / 450 VA', 'tegangan': 'TR'},
    {'group': 'I', 'golongan': 'I', 'sub_golongan': 'I.1 / 900 VA', 'tegangan': 'TR'},
    {'group': 'I', 'golongan': 'I', 'sub_golongan': 'I.1 / 1.300 VA', 'tegangan': 'TR'},
    {'group': 'I', 'golongan': 'I', 'sub_golongan': 'I.1 / 2.200 VA', 'tegangan': 'TR'},
    {'group': 'I', 'golongan': 'I', 'sub_golongan': 'I.1 / 3.500 s.d 14 kVA', 'tegangan': 'TR'},
    {'group': 'I', 'golongan': 'I', 'sub_golongan': 'I.2 / > 14 kVA s.d 200 kVA', 'tegangan': 'TR'},
    {'group': 'I', 'golongan': 'I', 'sub_golongan': 'I.3 / > 200 kVA', 'tegangan': 'TM'},
    {'group': 'I', 'golongan': 'I', 'sub_golongan': 'I.4 / 30.000 kVA keatas', 'tegangan': 'TT'},
    {'group': 'P', 'golongan': 'P', 'sub_golongan': 'P.1 / 450 VA', 'tegangan': 'TR'},
    {'group': 'P', 'golongan': 'P', 'sub_golongan': 'P.1 / 900 VA', 'tegangan': 'TR'},
    {'group': 'P', 'golongan': 'P', 'sub_golongan': 'P.1 / 1.300 VA', 'tegangan': 'TR'},
    {'group': 'P', 'golongan': 'P', 'sub_golongan': 'P.1 / 2.200 VA s.d 5.500 VA', 'tegangan': 'TR'},
    {'group': 'P', 'golongan': 'P', 'sub_golongan': 'P.1 / 6.600 VA s.d 200 kVA', 'tegangan': 'TR'},
    {'group': 'P', 'golongan': 'P', 'sub_golongan': 'P.2 / > 200 kVA', 'tegangan': 'TM'},
    {'group': 'P', 'golongan': 'P', 'sub_golongan': 'P.3 (khusus)', 'tegangan': 'TT'},
    {'group': 'TCL', 'golongan': 'T', 'sub_golongan': 'T / TM > 200 kVA s.d < 30.000 kVA', 'tegangan': 'TM'},
    {'group': 'TCL', 'golongan': 'T', 'sub_golongan': 'T / TT 30.000 kVA keatas', 'tegangan': 'TT'},
    {'group': 'TCL', 'golongan': 'C', 'sub_golongan': 'C / TR s.d 200 kVA', 'tegangan': 'TR'},
    {'group': 'TCL', 'golongan': 'C', 'sub_golongan': 'C / TM > 200 kVA s.d < 30.000 kVA', 'tegangan': 'TM'},
    {'group': 'TCL', 'golongan': 'C', 'sub_golongan': 'C / TT 30.000 kVA keatas', 'tegangan': 'TT'},
    {'group': 'TCL', 'golongan': 'L', 'sub_golongan': 'L / TR s.d 200 kVA', 'tegangan': 'TR'},
    {'group': 'TCL', 'golongan': 'L', 'sub_golongan': 'L / TM > 200 kVA s.d < 30.000 kVA', 'tegangan': 'TM'},
    {'group': 'TCL', 'golongan': 'L', 'sub_golongan': 'L / TT 30.000 kVA keatas', 'tegangan': 'TT'},
]
KWH_JUAL_SUB_INDEX = {item['sub_golongan']: item for item in KWH_JUAL_CATALOG}


def _json_error(message, status=400):
    return jsonify({'error': message}), status


def _wants_json():
    return request.path.startswith('/api/') or request.accept_mimetypes.best == 'application/json'


def _client_ip():
    forwarded = request.headers.get('X-Forwarded-For')
    if forwarded:
        return forwarded.split(',')[0].strip()
    return request.remote_addr or 'unknown'


def _current_user():
    user_id = session.get('user_id')
    if not user_id:
        return None
    return db.session.get(User, user_id)


def _ensure_csrf_token():
    token = session.get('csrf_token')
    if not token:
        token = secrets.token_urlsafe(32)
        session['csrf_token'] = token
    return token


def csrf_token():
    return _ensure_csrf_token()


def _validate_csrf():
    expected = session.get('csrf_token')
    supplied = (
        request.headers.get('X-CSRFToken') or
        request.headers.get('X-CSRF-Token') or
        request.form.get('csrf_token')
    )
    return bool(expected and supplied and secrets.compare_digest(expected, supplied))


def _login_user(user):
    session.clear()
    session.permanent = True
    session['user_id'] = user.id
    session['username'] = user.username
    session['role'] = user.role
    _ensure_csrf_token()
    user.last_login_at = datetime.utcnow()


def _logout_user():
    session.clear()


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not getattr(g, 'current_user', None):
            if _wants_json():
                return _json_error('Login diperlukan.', 401)
            return redirect(url_for('login', next=request.full_path if request.query_string else request.path))
        return view(*args, **kwargs)
    return wrapped


def role_required(*roles):
    def decorator(view):
        @wraps(view)
        def wrapped(*args, **kwargs):
            user = getattr(g, 'current_user', None)
            if not user:
                if _wants_json():
                    return _json_error('Login diperlukan.', 401)
                return redirect(url_for('login', next=request.path))
            if user.role not in roles:
                if _wants_json():
                    return _json_error('Akses ditolak untuk role ini.', 403)
                abort(403)
            return view(*args, **kwargs)
        return wrapped
    return decorator


def _prune_events(events, window_minutes):
    cutoff = datetime.utcnow() - timedelta(minutes=window_minutes)
    while events and events[0] < cutoff:
        events.popleft()


def _rate_limited(bucket, key, limit, window_minutes):
    events = bucket[key]
    _prune_events(events, window_minutes)
    return len(events) >= limit


def _record_rate_event(bucket, key, window_minutes):
    events = bucket[key]
    _prune_events(events, window_minutes)
    events.append(datetime.utcnow())


def _clear_rate_events(bucket, key):
    bucket.pop(key, None)


def _is_login_locked(key):
    until = LOGIN_LOCKOUTS.get(key)
    if not until:
        return False
    if until <= datetime.utcnow():
        LOGIN_LOCKOUTS.pop(key, None)
        return False
    return True


def _lock_login(key):
    LOGIN_LOCKOUTS[key] = datetime.utcnow() + timedelta(minutes=app.config.get('LOGIN_LOCKOUT_MINUTES', 15))


def _login_rate_key(username):
    return f'{_client_ip()}:{(username or "").lower()}'


def _upload_rate_key():
    user = getattr(g, 'current_user', None)
    return f'{user.id if user else "anon"}:{_client_ip()}'


def _validate_password_policy(password):
    min_len = app.config.get('PASSWORD_MIN_LENGTH', 10)
    if len(password or '') < min_len:
        return f'Password minimal {min_len} karakter.'
    checks = [
        any(ch.islower() for ch in password),
        any(ch.isupper() for ch in password),
        any(ch.isdigit() for ch in password),
        any(not ch.isalnum() for ch in password),
    ]
    if sum(checks) < 3:
        return 'Password harus memakai minimal 3 jenis karakter: huruf kecil, huruf besar, angka, atau simbol.'
    return None


def _request_payload():
    if request.is_json:
        return request.get_json(silent=True) or {}
    return request.form.to_dict()


def _bool_value(value):
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {'1', 'true', 'yes', 'on', 'aktif'}


def _normalize_workflow_status(value):
    raw = str(value or '').strip().upper().replace('-', '_').replace(' ', '_')
    compact = raw.replace('_', '')
    aliases = {
        'DRAFT': 'DRAFT',
        'SUDAHUPLOAD': 'SUDAH_UPLOAD',
        'UPLOAD': 'SUDAH_UPLOAD',
        'SUDAHDICEK': 'SUDAH_DICEK',
        'DICEK': 'SUDAH_DICEK',
        'CHECKED': 'SUDAH_DICEK',
        'FINAL': 'FINAL',
        'TERKUNCI': 'TERKUNCI',
        'LOCKED': 'TERKUNCI',
    }
    status = aliases.get(compact, raw)
    if status not in WORKFLOW_STATUS_ORDER:
        raise ValueError('Status workflow tidak dikenali.')
    return status


def _workflow_period(value):
    try:
        return _month_date(value)
    except ValueError:
        raise ValueError('Format periode harus YYYY-MM.')


def _workflow_record(period, create=False):
    record = MonthlyDataStatus.query.filter_by(periode_bulan=period).first()
    if not record and create:
        record = MonthlyDataStatus(periode_bulan=period, status='DRAFT')
        db.session.add(record)
        db.session.flush()
    return record


def _workflow_allowed_statuses(status, user=None):
    allowed = list(WORKFLOW_TRANSITIONS.get(status, ['DRAFT']))
    if status == 'TERKUNCI' and (not user or user.role != 'admin'):
        return ['TERKUNCI']
    if 'TERKUNCI' in allowed and (not user or user.role != 'admin'):
        allowed.remove('TERKUNCI')
    return allowed


def _workflow_payload(period, record=None):
    record = record if record is not None else _workflow_record(period)
    status = _normalize_workflow_status(record.status) if record else 'DRAFT'
    current_index = WORKFLOW_STATUS_ORDER.index(status)
    user = getattr(g, 'current_user', None)
    allowed = _workflow_allowed_statuses(status, user)
    return {
        'id': record.id if record else None,
        'periode': period.strftime('%Y-%m'),
        'periode_bulan': period.strftime('%Y-%m-%d'),
        'status': status,
        'label': WORKFLOW_STATUS_LABELS[status],
        'catatan': record.catatan if record else '',
        'locked': status == 'TERKUNCI',
        'writable': status in WORKFLOW_WRITABLE_STATUSES,
        'locked_at': record.locked_at.isoformat() if record and record.locked_at else None,
        'locked_by': record.locked_by if record else None,
        'updated_at': record.updated_at.isoformat() if record and record.updated_at else None,
        'allowed_next': [
            {'status': code, 'label': WORKFLOW_STATUS_LABELS[code]}
            for code in allowed
        ],
        'steps': [
            {
                'status': code,
                'label': WORKFLOW_STATUS_LABELS[code],
                'done': index < current_index,
                'active': code == status,
                'locked': code == 'TERKUNCI',
            }
            for index, code in enumerate(WORKFLOW_STATUS_ORDER)
        ],
    }


def _ensure_period_writable(period):
    record = _workflow_record(period)
    if not record:
        return
    status = _normalize_workflow_status(record.status)
    if status not in WORKFLOW_WRITABLE_STATUSES:
        label = WORKFLOW_STATUS_LABELS[status]
        raise ValueError(
            f'Periode {period.strftime("%Y-%m")} berstatus {label}. '
            'Turunkan status ke Draft/Sudah Upload sebelum import ulang.'
        )


def _mark_period_uploaded(period, source, filename=None):
    record = _workflow_record(period, create=True)
    status = _normalize_workflow_status(record.status)
    if status == 'DRAFT':
        record.status = 'SUDAH_UPLOAD'
    if not record.catatan:
        record.catatan = f'Upload terakhir dari {source}.'
    record.locked_at = None if record.status != 'TERKUNCI' else record.locked_at
    record.locked_by = None if record.status != 'TERKUNCI' else record.locked_by
    _audit('MARK_MONTH_UPLOADED', entity_type='monthly_data_status', entity_id=record.id, detail={
        'periode_bulan': period.strftime('%Y-%m-%d'),
        'source': source,
        'filename': filename,
        'status': record.status,
    })
    return record


def _audit_detail(record):
    try:
        return json.loads(record.detail_json or '{}')
    except (TypeError, ValueError):
        return {}


def _audit_month_summary(detail):
    labels = {
        'filename': 'File',
        'source': 'Sumber',
        'from_status': 'Dari',
        'to_status': 'Ke',
        'created': 'Baru',
        'updated': 'Update',
        'alerts': 'Alert',
        'error_count': 'Error',
        'feeder_count': 'Penyulang',
        'gi_count': 'GI',
    }
    parts = []
    for key, label in labels.items():
        value = detail.get(key)
        if value in (None, '', []):
            continue
        parts.append(f'{label}: {value}')
    return '; '.join(parts) or '-'


def _monthly_activity_payload(period, limit=30):
    period_day = period.strftime('%Y-%m-%d')
    period_month = period.strftime('%Y-%m')
    rows = AuditLog.query.filter(AuditLog.action.in_(MONTHLY_ACTIVITY_ACTIONS)).filter(
        (AuditLog.detail_json.contains(period_day)) |
        (AuditLog.detail_json.contains(period_month))
    ).order_by(AuditLog.created_at.desc()).limit(limit).all()
    activities = []
    for row in rows:
        detail = _audit_detail(row)
        activities.append({
            'id': row.id,
            'created_at': row.created_at.isoformat() if row.created_at else None,
            'username': row.username or '-',
            'role': row.role or '-',
            'action': row.action,
            'status': row.status,
            'summary': _audit_month_summary(detail),
            'detail': detail,
        })
    return {
        'periode': period_month,
        'periode_bulan': period_day,
        'rows': activities,
    }


def _readiness_status(value, expected=None, optional=False):
    if expected and expected > 0:
        ratio = min(float(value or 0) / float(expected), 1)
    else:
        ratio = 1 if value else 0
    if ratio >= .98:
        return 'ready', ratio
    if ratio > 0:
        return 'partial', ratio
    return ('optional' if optional else 'empty'), ratio


def _readiness_item(code, label, value, expected=None, optional=False, detail=''):
    status, ratio = _readiness_status(value, expected, optional)
    if expected and expected > 0:
        subtitle = f'{int(value or 0)} dari {int(expected)}'
    else:
        subtitle = f'{int(value or 0)} data'
    return {
        'code': code,
        'label': label,
        'value': int(value or 0),
        'expected': int(expected) if expected is not None else None,
        'ratio': round(ratio, 4),
        'percent': round(ratio * 100),
        'status': status,
        'optional': optional,
        'subtitle': subtitle,
        'detail': detail,
    }


def _readiness_payload(period):
    active_gi = GarduInduk.query.filter_by(aktif=True).count()
    active_trafo = Trafo.query.filter_by(aktif=True).count()
    active_feeders = Penyulang.query.filter_by(aktif=True).count()

    feeder_rows = FeederReading.query.filter_by(periode_bulan=period).count()
    feeder_unique = db.session.query(func.count(func.distinct(FeederReading.penyulang_id))).filter(
        FeederReading.periode_bulan == period
    ).scalar() or 0
    alert_count = FeederReading.query.filter_by(
        periode_bulan=period,
        flag_alert=True,
    ).count()

    mu_total_expr = (
        func.coalesce(MeterReading.mu_kwh_wbp, 0) +
        func.coalesce(MeterReading.mu_kwh_lwbp1, 0) +
        func.coalesce(MeterReading.mu_kwh_lwbp2, 0)
    )
    mp_total_expr = (
        func.coalesce(MeterReading.mp_kwh_wbp, 0) +
        func.coalesce(MeterReading.mp_kwh_lwbp1, 0) +
        func.coalesce(MeterReading.mp_kwh_lwbp2, 0)
    )
    mu_count = MeterReading.query.filter(
        MeterReading.periode_bulan == period,
        mu_total_expr > 0,
    ).count()
    mp_count = MeterReading.query.filter(
        MeterReading.periode_bulan == period,
        mp_total_expr > 0,
    ).count()
    exim_count = EximMonthlyResult.query.filter_by(periode_bulan=period).count()
    transfer_count = TransferAntarUnit.query.filter_by(periode_bulan=period).count()
    rekap_count = RekapBulanan.query.filter_by(periode_bulan=period).count()

    items = [
        _readiness_item('master_gi', 'Master GI', active_gi, detail='gardu induk aktif'),
        _readiness_item('master_trafo', 'Master Trafo', active_trafo, detail='trafo aktif'),
        _readiness_item('master_penyulang', 'Master Penyulang', active_feeders, detail='penyulang aktif'),
        _readiness_item('feeder_reading', 'kWh Penyulang', feeder_unique, active_feeders, detail=f'{feeder_rows} baris pembacaan'),
        _readiness_item('meter_utama', 'kWh Utama', mu_count, active_trafo, detail='trafo punya meter utama'),
        _readiness_item('meter_pembanding', 'kWh Pembanding', mp_count, active_trafo, detail='trafo punya meter pembanding'),
        _readiness_item('exim', 'Transfer EXIM', exim_count, optional=True, detail='snapshot transfer EXIM'),
        _readiness_item('transfer_uid', 'Transfer Antar UID', transfer_count, optional=True, detail='transaksi antar UID'),
        _readiness_item('rekap', 'Rekap Bulanan', rekap_count, optional=True, detail='snapshot rekap'),
    ]
    required = [item for item in items if not item['optional']]
    score = round(sum(item['ratio'] for item in required) / len(required) * 100) if required else 0
    blockers = [
        item['label']
        for item in required
        if item['status'] in {'empty', 'partial'}
    ]
    return {
        'periode': period.strftime('%Y-%m'),
        'periode_bulan': period.strftime('%Y-%m-%d'),
        'score': score,
        'status': 'ready' if not blockers else 'partial' if score else 'empty',
        'can_finalize': not blockers,
        'blockers': blockers,
        'alert_count': alert_count,
        'items': items,
    }


def _next_month(period):
    return date(period.year + 1, 1, 1) if period.month == 12 else date(period.year, period.month + 1, 1)


def _report_period_bounds(default_month=False):
    bulan = (request.args.get('bulan') or request.args.get('periode') or '').strip()
    tahun = request.args.get('tahun', type=int) or date.today().year
    bulan_int = request.args.get('month', type=int)
    if bulan:
        period = _month_date(bulan)
        return period, _next_month(period), period.strftime('%Y-%m')
    if bulan_int:
        period = date(tahun, bulan_int, 1)
        return period, _next_month(period), period.strftime('%Y-%m')
    if default_month:
        period = date.today().replace(day=1)
        return period, _next_month(period), period.strftime('%Y-%m')
    return date(tahun, 1, 1), date(tahun + 1, 1, 1), str(tahun)


def _month_filter(query, column, start, end):
    return query.filter(column >= start, column < end)


def _kwh_sum(*columns):
    expr = 0
    for column in columns:
        if isinstance(column, type) and issubclass(column, db.Model):
            continue
        expr += func.coalesce(column, 0)
    return expr


def _float_value(value):
    return float(value or 0)


def _module_access_payload(role=None):
    role = (role or '').strip().lower()
    rows = []
    for item in MODULE_ACCESS_MATRIX:
        access = item['access']
        row = {
            'module': item['module'],
            'group': item['group'],
            'access': access,
        }
        if role in ROLES:
            row['role'] = role
            row['allowed_actions'] = [
                action
                for action, roles in access.items()
                if role in roles
            ]
        rows.append(row)
    return rows


def _report_dataset(module):
    module = module.replace('-', '_').lower()
    if module in {'rekap', 'rekap_kwh'}:
        return _report_rekap_kwh()
    if module in {'deviasi', 'deviasi_gi'}:
        return _report_deviasi_gi()
    if module == 'proporsional':
        return _report_proporsional()
    if module in {'transfer_exim', 'exim'}:
        return _report_transfer_exim()
    if module in {'transfer_uid', 'transfer_antar_uid'}:
        return _report_transfer_uid()
    raise ValueError('Modul export tidak dikenali.')


def _report_rekap_kwh():
    start, end, period_label = _report_period_bounds()
    q = db.session.query(RekapBulanan, GarduInduk, Trafo).join(
        GarduInduk, RekapBulanan.gi_id == GarduInduk.id
    ).outerjoin(Trafo, RekapBulanan.trafo_id == Trafo.id)
    q = _month_filter(q, RekapBulanan.periode_bulan, start, end)
    headers = ['Periode', 'Gardu Induk', 'Trafo', 'MU Total', 'MP Total', 'Penyulang Total', 'Dev MU-MP %', 'Dev MU-Penyulang %', 'Susut kWh', 'Susut %', 'Ekspor', 'Impor']
    rows = []
    for rekap, gi, trafo in q.order_by(RekapBulanan.periode_bulan, GarduInduk.nama_gi, Trafo.kode_trafo).all():
        rows.append([
            rekap.periode_bulan.strftime('%Y-%m'),
            gi.nama_gi,
            trafo.nama_trafo if trafo else 'TOTAL GI',
            _float_value(rekap.kwh_mu_total),
            _float_value(rekap.kwh_mp_total),
            _float_value(rekap.kwh_penyulang_total),
            _float_value(rekap.deviasi_mu_mp),
            _float_value(rekap.deviasi_mu_penyulang),
            _float_value(rekap.susut_kwh),
            _float_value(rekap.susut_persen),
            _float_value(rekap.transfer_ekspor),
            _float_value(rekap.transfer_impor),
        ])
    return 'Rekap kWh', f'Periode {period_label}', headers, rows, f'rekap_kwh_{period_label}'


def _report_deviasi_gi():
    start, end, period_label = _report_period_bounds()
    gi_id = request.args.get('gi_id', type=int)
    meter_expr = _kwh_sum(MeterReading, MeterReading.mu_kwh_wbp, MeterReading.mu_kwh_lwbp1, MeterReading.mu_kwh_lwbp2)
    mp_expr = _kwh_sum(MeterReading, MeterReading.mp_kwh_wbp, MeterReading.mp_kwh_lwbp1, MeterReading.mp_kwh_lwbp2)
    feeder_expr = _kwh_sum(FeederReading, FeederReading.kwh_wbp, FeederReading.kwh_lwbp1, FeederReading.kwh_lwbp2)
    feeder_sub = db.session.query(
        FeederReading.trafo_id,
        FeederReading.periode_bulan,
        func.sum(feeder_expr).label('feeder_total'),
    ).filter(
        FeederReading.periode_bulan >= start,
        FeederReading.periode_bulan < end,
    ).group_by(FeederReading.trafo_id, FeederReading.periode_bulan).subquery()
    q = db.session.query(
        MeterReading.periode_bulan,
        GarduInduk.nama_gi,
        Trafo.kode_trafo,
        Trafo.nama_trafo,
        meter_expr.label('mu_total'),
        mp_expr.label('mp_total'),
        func.coalesce(feeder_sub.c.feeder_total, 0).label('feeder_total'),
    ).join(Trafo, MeterReading.trafo_id == Trafo.id).join(
        GarduInduk, MeterReading.gi_id == GarduInduk.id
    ).outerjoin(
        feeder_sub,
        (feeder_sub.c.trafo_id == MeterReading.trafo_id) &
        (feeder_sub.c.periode_bulan == MeterReading.periode_bulan),
    )
    q = _month_filter(q, MeterReading.periode_bulan, start, end)
    if gi_id:
        q = q.filter(MeterReading.gi_id == gi_id)
    headers = ['Periode', 'Gardu Induk', 'Kode Trafo', 'Trafo', 'MU Total', 'MP Total', 'Penyulang Total', 'Dev MU-Penyulang kWh', 'Dev MU-Penyulang %', 'Dev MU-MP %']
    rows = []
    for row in q.order_by(MeterReading.periode_bulan, GarduInduk.nama_gi, Trafo.kode_trafo).all():
        mu = _float_value(row.mu_total)
        mp = _float_value(row.mp_total)
        feeder = _float_value(row.feeder_total)
        gap = mu - feeder
        rows.append([
            row.periode_bulan.strftime('%Y-%m'),
            row.nama_gi,
            row.kode_trafo,
            row.nama_trafo,
            mu,
            mp,
            feeder,
            gap,
            (gap / mu * 100) if mu else 0,
            ((mu - mp) / mu * 100) if mu else 0,
        ])
    return 'Deviasi GI', f'Periode {period_label}', headers, rows, f'deviasi_gi_{period_label}'


def _report_proporsional():
    start, end, period_label = _report_period_bounds(default_month=True)
    gi_id = request.args.get('gi_id', type=int)
    mu_expr = _kwh_sum(MeterReading, MeterReading.mu_kwh_wbp, MeterReading.mu_kwh_lwbp1, MeterReading.mu_kwh_lwbp2)
    feeder_expr = _kwh_sum(FeederReading, FeederReading.kwh_wbp, FeederReading.kwh_lwbp1, FeederReading.kwh_lwbp2)
    mu_rows = db.session.query(MeterReading.gi_id, func.sum(mu_expr).label('mu_total')).filter(
        MeterReading.periode_bulan >= start,
        MeterReading.periode_bulan < end,
    )
    if gi_id:
        mu_rows = mu_rows.filter(MeterReading.gi_id == gi_id)
    mu_by_gi = {row.gi_id: _float_value(row.mu_total) for row in mu_rows.group_by(MeterReading.gi_id).all()}
    q = db.session.query(FeederReading, Penyulang, Trafo, GarduInduk).join(
        Penyulang, FeederReading.penyulang_id == Penyulang.id
    ).join(Trafo, FeederReading.trafo_id == Trafo.id).join(
        GarduInduk, FeederReading.gi_id == GarduInduk.id
    )
    q = _month_filter(q, FeederReading.periode_bulan, start, end)
    if gi_id:
        q = q.filter(FeederReading.gi_id == gi_id)
    raw_rows = q.order_by(GarduInduk.nama_gi, Trafo.kode_trafo, Penyulang.kode_penyulang).all()
    feeder_total_by_gi = defaultdict(float)
    for reading, penyulang, trafo, gi in raw_rows:
        feeder_total_by_gi[gi.id] += reading.kwh_total
    headers = ['Periode', 'Gardu Induk', 'Trafo', 'Area/UP3', 'Penyulang', 'Hasil Baca', 'Porsi %', 'Deviasi Dibagi', 'Total Proporsional']
    rows = []
    for reading, penyulang, trafo, gi in raw_rows:
        total_feeder_gi = feeder_total_by_gi.get(gi.id, 0)
        mu_total = mu_by_gi.get(gi.id, 0)
        deviasi = mu_total - total_feeder_gi
        porsi = (reading.kwh_total / total_feeder_gi) if total_feeder_gi else 0
        alokasi = deviasi * porsi
        rows.append([
            reading.periode_bulan.strftime('%Y-%m'),
            gi.nama_gi,
            trafo.kode_trafo,
            penyulang.area_up3 or 'Belum Dipetakan',
            penyulang.nama_penyulang,
            reading.kwh_total,
            porsi * 100,
            alokasi,
            reading.kwh_total + alokasi,
        ])
    return 'Proporsional', f'Periode {period_label}', headers, rows, f'proporsional_{period_label}'


def _report_transfer_exim():
    start, end, period_label = _report_period_bounds()
    q = db.session.query(EximMonthlyResult, EximRule).join(EximRule, EximMonthlyResult.rule_id == EximRule.id)
    q = _month_filter(q, EximMonthlyResult.periode_bulan, start, end)
    headers = ['Periode', 'Rule', 'Metode', 'UP3 Asal', 'UP3 Tujuan', 'Fungsi', 'Arah', 'Basis kWh', 'WBP', 'LWBP1', 'LWBP2', 'Transfer kWh', 'Porsi %', 'Catatan']
    rows = []
    for result, rule in q.order_by(EximMonthlyResult.periode_bulan, EximMonthlyResult.up3_tujuan).all():
        rows.append([
            result.periode_bulan.strftime('%Y-%m'),
            rule.nama_rule or rule.kode_rule,
            result.metode,
            result.up3_asal,
            result.up3_tujuan,
            result.fungsi,
            result.arah,
            _float_value(result.kwh_basis),
            _float_value(result.kwh_wbp),
            _float_value(result.kwh_lwbp1),
            _float_value(result.kwh_lwbp2),
            _float_value(result.kwh_transfer),
            _float_value(result.porsi) * 100 if result.porsi is not None else 0,
            result.catatan,
        ])
    return 'Transfer EXIM', f'Periode {period_label}', headers, rows, f'transfer_exim_{period_label}'


def _report_transfer_uid():
    start, end, period_label = _report_period_bounds()
    q = _month_filter(TransferAntarUnit.query, TransferAntarUnit.periode_bulan, start, end)
    headers = ['Periode', 'Unit Asal', 'Unit Tujuan', 'GI/Interkoneksi', 'Kode Interbus', 'Arah', 'Transfer kWh', 'Catatan']
    rows = []
    for item in q.order_by(TransferAntarUnit.periode_bulan, TransferAntarUnit.unit_asal).all():
        rows.append([
            item.periode_bulan.strftime('%Y-%m'),
            item.unit_asal,
            item.unit_tujuan,
            item.gi_interkoneksi,
            item.kode_interbus,
            item.arah,
            _float_value(item.kwh_transfer),
            item.catatan,
        ])
    return 'Transfer Antar UID', f'Periode {period_label}', headers, rows, f'transfer_antar_uid_{period_label}'


def _report_filter_label():
    labels = []
    for key, label in [
        ('tahun', 'Tahun'),
        ('month', 'Bulan'),
        ('bulan', 'Periode'),
        ('periode', 'Periode'),
        ('gi_id', 'GI ID'),
    ]:
        value = (request.args.get(key) or '').strip()
        if value:
            labels.append(f'{label}: {value}')
    return ' | '.join(labels) if labels else 'Semua data'


def _sumable_report_column(header):
    text = str(header or '').lower()
    if '%' in text or 'porsi' in text or 'faktor' in text:
        return False
    return any(token in text for token in [
        'kwh', 'wbp', 'lwbp', 'mu', 'mp', 'penyulang', 'susut',
        'transfer', 'ekspor', 'impor', 'hasil baca', 'proporsional',
        'deviasi dibagi', 'basis',
    ])


def _excel_report(title, subtitle, headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Laporan'
    ws.append([title])
    ws.append([subtitle])
    ws.append([f'Dibuat: {datetime.now().strftime("%d/%m/%Y %H:%M")} | Filter: {_report_filter_label()}'])
    ws.append([])
    ws.append(headers)
    for row in rows:
        ws.append(row)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(headers), 1))
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max(len(headers), 1))
    ws['A1'].font = Font(size=14, bold=True)
    ws['A2'].font = Font(size=10, italic=True, color='667085')
    ws['A3'].font = Font(size=9, color='667085')
    header_row = 5
    fill = PatternFill('solid', fgColor='E8F1FF')
    total_fill = PatternFill('solid', fgColor='F3F6FB')
    thin = Side(style='thin', color='D7DEE8')
    border = Border(top=thin, right=thin, bottom=thin, left=thin)
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color='1D2430')
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    first_data_row = header_row + 1
    last_data_row = ws.max_row
    for row_idx in range(first_data_row, last_data_row + 1):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=not isinstance(cell.value, (int, float)))
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00' if ('%' in str(headers[col_idx - 1]) or abs(cell.value) % 1) else '#,##0'

    if rows:
        total_row = ws.max_row + 1
        ws.cell(total_row, 1, 'TOTAL')
        for col_idx, header in enumerate(headers, start=1):
            values = [
                row[col_idx - 1]
                for row in rows
                if len(row) >= col_idx and isinstance(row[col_idx - 1], (int, float))
            ]
            if values and _sumable_report_column(header):
                ws.cell(total_row, col_idx, sum(values))
        for cell in ws[total_row]:
            cell.font = Font(bold=True, color='1D2430')
            cell.fill = total_fill
            cell.border = border
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00' if abs(cell.value) % 1 else '#,##0'

    ws.freeze_panes = 'A6'
    ws.auto_filter.ref = f'A{header_row}:{get_column_letter(max(len(headers), 1))}{max(ws.max_row, header_row)}'
    ws.print_title_rows = f'{header_row}:{header_row}'
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    for col_idx in range(1, len(headers) + 1):
        width = max(
            len(str(ws.cell(row=row_idx, column=col_idx).value or ''))
            for row_idx in range(1, min(ws.max_row, 80) + 1)
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 12), 34)
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def _pdf_escape(value):
    return str(value if value is not None else '').replace('\\', '\\\\').replace('(', '\\(').replace(')', '\\)')


def _pdf_report(title, subtitle, headers, rows):
    text_rows = [' | '.join(headers)]
    text_rows.extend(' | '.join(str(value if value is not None else '') for value in row) for row in rows[:120])
    lines = [
        title,
        subtitle,
        f'Dibuat: {datetime.now().strftime("%d/%m/%Y %H:%M")} | Filter: {_report_filter_label()}',
        f'Total baris: {len(rows)}',
        '',
    ] + [line[:145] for line in text_rows]
    pages = [lines[i:i + 42] for i in range(0, len(lines), 42)] or [[]]
    objects = [
        b'<< /Type /Catalog /Pages 2 0 R >>',
        None,
        b'<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>',
    ]
    page_ids = []
    for page_lines in pages:
        content_lines = ['BT', '/F1 9 Tf']
        y = 800
        for line in page_lines:
            content_lines.append(f'1 0 0 1 36 {y} Tm ({_pdf_escape(line)}) Tj')
            y -= 17
        content_lines.append('ET')
        stream = '\n'.join(content_lines).encode('latin-1', 'replace')
        page_id = len(objects) + 1
        content_id = len(objects) + 2
        page_ids.append(page_id)
        objects.append(f'<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Resources << /Font << /F1 3 0 R >> >> /Contents {content_id} 0 R >>'.encode('latin-1'))
        objects.append(b'<< /Length ' + str(len(stream)).encode('ascii') + b' >>\nstream\n' + stream + b'\nendstream')
    kids = ' '.join(f'{page_id} 0 R' for page_id in page_ids)
    objects[1] = f'<< /Type /Pages /Kids [{kids}] /Count {len(page_ids)} >>'.encode('latin-1')
    output = io.BytesIO()
    output.write(b'%PDF-1.4\n')
    offsets = [0]
    for idx, obj in enumerate(objects, start=1):
        offsets.append(output.tell())
        output.write(f'{idx} 0 obj\n'.encode('ascii'))
        output.write(obj)
        output.write(b'\nendobj\n')
    xref_pos = output.tell()
    output.write(f'xref\n0 {len(objects) + 1}\n'.encode('ascii'))
    output.write(b'0000000000 65535 f \n')
    for offset in offsets[1:]:
        output.write(f'{offset:010d} 00000 n \n'.encode('ascii'))
    output.write(f'trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_pos}\n%%EOF'.encode('ascii'))
    return output.getvalue()


def _report_file_response(module, fmt):
    title, subtitle, headers, rows, filename = _report_dataset(module)
    fmt = fmt.lower()
    if fmt not in {'xlsx', 'pdf'}:
        raise ValueError('Format export harus xlsx atau pdf.')
    _audit('EXPORT_REPORT', entity_type='report', entity_id=module, detail={
        'module': module,
        'format': fmt,
        'rows': len(rows),
        'subtitle': subtitle,
    })
    db.session.commit()
    if fmt == 'xlsx':
        stream = _excel_report(title, subtitle, headers, rows)
        return send_file(
            stream,
            as_attachment=True,
            download_name=f'{filename}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    if fmt == 'pdf':
        pdf_bytes = _pdf_report(title, subtitle, headers, rows)
        return Response(
            pdf_bytes,
            mimetype='application/pdf',
            headers={'Content-Disposition': f'attachment; filename="{filename}.pdf"'},
        )


def _check_upload_rate():
    key = _upload_rate_key()
    if _rate_limited(
        UPLOAD_EVENTS,
        key,
        app.config.get('UPLOAD_RATE_LIMIT', 10),
        app.config.get('UPLOAD_RATE_WINDOW_MINUTES', 10),
    ):
        raise ValueError('Terlalu banyak upload dalam waktu singkat. Coba lagi beberapa menit lagi.')
    _record_rate_event(UPLOAD_EVENTS, key, app.config.get('UPLOAD_RATE_WINDOW_MINUTES', 10))


def _audit(action, entity_type=None, entity_id=None, detail=None, status='SUCCESS', username=None):
    user = getattr(g, 'current_user', None)
    record = AuditLog(
        user_id=user.id if user else None,
        username=username or (user.username if user else None),
        role=user.role if user else None,
        action=action,
        entity_type=entity_type,
        entity_id=str(entity_id) if entity_id is not None else None,
        status=status,
        ip_address=_client_ip(),
        user_agent=(request.headers.get('User-Agent') or '')[:255],
        detail_json=json.dumps(detail or {}, ensure_ascii=False),
    )
    db.session.add(record)


def _safe_commit_audit(action, detail=None, status='SUCCESS', username=None):
    try:
        _audit(action, detail=detail, status=status, username=username)
        db.session.commit()
    except Exception:
        db.session.rollback()


def _extension(filename):
    return secure_filename(filename or '').rsplit('.', 1)[-1].lower() if '.' in (filename or '') else ''


def _validate_upload_file(file, allowed_extensions):
    filename = secure_filename(file.filename or '')
    ext = _extension(filename)
    if not filename:
        raise ValueError('Nama file kosong.')
    if ext not in allowed_extensions:
        allowed = ', '.join(sorted(allowed_extensions))
        raise ValueError(f'Format file tidak diizinkan. Gunakan: {allowed}.')
    if request.content_length and request.content_length > app.config['MAX_CONTENT_LENGTH']:
        max_mb = app.config['MAX_CONTENT_LENGTH'] // (1024 * 1024)
        raise ValueError(f'Ukuran file melebihi batas {max_mb} MB.')

    pos = file.stream.tell()
    head = file.stream.read(8)
    file.stream.seek(pos)
    if ext in {'xlsx', 'xlsm'} and not head.startswith(b'PK'):
        raise ValueError('File Excel tidak valid atau rusak.')
    if ext == 'xls' and head != XLS_SIGNATURE:
        raise ValueError('File XLS tidak valid atau rusak.')
    return filename, ext


@app.context_processor
def inject_security_context():
    return {
        'csrf_token': csrf_token,
        'current_user': lambda: getattr(g, 'current_user', None),
    }


@app.before_request
def apply_security_gate():
    g.current_user = _current_user()
    if g.current_user and not g.current_user.aktif:
        _logout_user()
        g.current_user = None

    if request.endpoint in PUBLIC_ENDPOINTS:
        return None

    if request.method not in SAFE_METHODS and not _validate_csrf():
        if _wants_json():
            return _json_error('CSRF token tidak valid.', 403)
        abort(403)

    if app.config.get('SECURITY_REQUIRE_LOGIN', True) and not g.current_user:
        if _wants_json():
            return _json_error('Login diperlukan.', 401)
        return redirect(url_for('login', next=request.full_path if request.query_string else request.path))
    return None


@app.after_request
def add_security_headers(response):
    response.headers.setdefault('X-Content-Type-Options', 'nosniff')
    response.headers.setdefault('X-Frame-Options', 'SAMEORIGIN')
    response.headers.setdefault('Referrer-Policy', 'same-origin')
    response.headers.setdefault('Permissions-Policy', 'camera=(), microphone=(), geolocation=()')
    response.headers.setdefault(
        'Content-Security-Policy',
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' https://cdnjs.cloudflare.com; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com https://cdn.jsdelivr.net; "
        "font-src 'self' https://fonts.gstatic.com https://cdn.jsdelivr.net; "
        "img-src 'self' data:; connect-src 'self'; frame-ancestors 'self';"
    )
    return response


@app.cli.command('create-admin')
@click.option('--username', prompt=True)
@click.option('--password', prompt=True, hide_input=True, confirmation_prompt=True)
@click.option('--name', default='Administrator')
def create_admin_command(username, password, name):
    password_error = _validate_password_policy(password)
    if password_error:
        raise click.ClickException(password_error)
    existing = User.query.filter_by(username=username.strip()).first()
    if existing:
        existing.role = 'admin'
        existing.aktif = True
        existing.nama_lengkap = name or existing.nama_lengkap
        existing.set_password(password)
        action = 'diperbarui'
    else:
        user = User(username=username.strip(), nama_lengkap=name, role='admin', aktif=True)
        user.set_password(password)
        db.session.add(user)
        action = 'dibuat'
    db.session.commit()
    click.echo(f'Admin {username} berhasil {action}.')


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        if not _validate_csrf():
            abort(403)
        username = (request.form.get('username') or '').strip()
        password = request.form.get('password') or ''
        rate_key = _login_rate_key(username)
        if _is_login_locked(rate_key):
            _safe_commit_audit('LOGIN_RATE_LIMITED', detail={'username': username}, status='FAILED', username=username)
            return render_template('login.html',
                                   error='Terlalu banyak percobaan login gagal. Coba lagi beberapa menit lagi.',
                                   next_url=request.form.get('next', '')), 429
        if _rate_limited(
            LOGIN_FAILURES,
            rate_key,
            app.config.get('LOGIN_RATE_LIMIT', 5),
            app.config.get('LOGIN_RATE_WINDOW_MINUTES', 15),
        ):
            _safe_commit_audit('LOGIN_RATE_LIMITED', detail={'username': username}, status='FAILED', username=username)
            return render_template('login.html',
                                   error='Terlalu banyak percobaan login gagal. Coba lagi beberapa menit lagi.',
                                   next_url=request.form.get('next', '')), 429
        user = User.query.filter_by(username=username).first()
        if user and user.aktif and user.check_password(password):
            _login_user(user)
            g.current_user = user
            _clear_rate_events(LOGIN_FAILURES, rate_key)
            _audit('LOGIN', entity_type='user', entity_id=user.id, detail={'username': user.username})
            db.session.commit()
            next_url = request.form.get('next') or url_for('dashboard')
            if not next_url.startswith('/'):
                next_url = url_for('dashboard')
            return redirect(next_url)

        _record_rate_event(LOGIN_FAILURES, rate_key, app.config.get('LOGIN_RATE_WINDOW_MINUTES', 15))
        if _rate_limited(
            LOGIN_FAILURES,
            rate_key,
            app.config.get('LOGIN_RATE_LIMIT', 5),
            app.config.get('LOGIN_RATE_WINDOW_MINUTES', 15),
        ):
            _lock_login(rate_key)
        _safe_commit_audit('LOGIN_FAILED', detail={'username': username}, status='FAILED', username=username)
        return render_template('login.html', error='Username atau password tidak sesuai.', next_url=request.form.get('next', ''))

    if getattr(g, 'current_user', None):
        return redirect(url_for('dashboard'))
    return render_template('login.html', next_url=request.args.get('next', ''))


@app.route('/logout', methods=['POST'])
def logout():
    if not _validate_csrf():
        abort(403)
    username = session.get('username')
    _safe_commit_audit('LOGOUT', detail={'username': username}, username=username)
    _logout_user()
    return redirect(url_for('login'))


@app.route('/')
def dashboard():
    return render_template('dashboard.html')

@app.route('/penyulang')
def halaman_penyulang():
    return render_template('penyulang.html',
        eyebrow='Gardu Induk', judul='kWh Penyulang',
        icon='plug', desc='Data pembacaan meter per penyulang.')


def _render_meter_page(mode='utama'):
    is_pembanding = mode == 'pembanding'
    return render_template('meter_gi.html',
        eyebrow='Gardu Induk',
        judul='kWh Pembanding' if is_pembanding else 'kWh Utama',
        icon='gauge' if is_pembanding else 'bolt',
        desc='Data kWh meter pembanding per trafo gardu induk.' if is_pembanding else 'Data kWh meter utama per trafo gardu induk.',
        meter_mode='pembanding' if is_pembanding else 'utama',
        primary_total_label='Total MP Tahun' if is_pembanding else 'Total MU Tahun',
        secondary_total_label='Total MU Tahun' if is_pembanding else 'Total MP Tahun',
        primary_meta='meter pembanding' if is_pembanding else 'meter utama',
        secondary_meta='meter utama' if is_pembanding else 'meter pembanding',
        primary_icon='gauge' if is_pembanding else 'bolt',
        secondary_icon='bolt' if is_pembanding else 'gauge',
        primary_name='Meter Pembanding' if is_pembanding else 'Meter Utama',
        secondary_name='Meter Utama' if is_pembanding else 'Meter Pembanding',
        primary_short='MP' if is_pembanding else 'MU',
        secondary_short='MU' if is_pembanding else 'MP',
        chart_title='Tren MP vs MU' if is_pembanding else 'Tren MU vs MP',
        focus_chart_title='Pemakaian MP per Trafo' if is_pembanding else 'Pemakaian MU per Trafo',
        table_title='Rekap kWh Pembanding per Trafo' if is_pembanding else 'Rekap kWh Utama per Trafo',
        export_name='kwh_pembanding' if is_pembanding else 'kwh_utama')


@app.route('/kwh-utama')
def halaman_kwh_utama():
    return _render_meter_page('utama')


@app.route('/kwh-pembanding')
def halaman_kwh_pembanding():
    return _render_meter_page('pembanding')


@app.route('/meter-gi')
def halaman_meter_gi():
    return _render_meter_page('utama')


@app.route('/psgi')
def halaman_psgi():
    return render_template('rekap.html',
        eyebrow='Gardu Induk', judul='PSGI',
        icon='building-factory-2',
        desc='Pemakaian sendiri gardu induk per periode dan relasinya terhadap perhitungan susut.')

@app.route('/deviasi')
def halaman_deviasi():
    return render_template('deviasi.html',
        eyebrow='Gardu Induk', judul='Deviasi',
        icon='chart-bar', desc='Perbandingan MU vs MP vs total penyulang.')

@app.route('/proporsional')
def halaman_proporsional():
    return render_template('proporsional.html',
        eyebrow='UID', judul='Proporsional',
        icon='percentage', desc='Alokasi energi proporsional per penyulang.')

@app.route('/transfer-antar-uid')
def halaman_transfer_antar_uid():
    return render_template('transfer_uid.html',
        eyebrow='UID', judul='Transfer Antar UID',
        icon='arrows-transfer-up', desc='Rekap ekspor dan impor energi antar UID.')

@app.route('/transfer')
def halaman_transfer():
    return render_template('transfer.html',
        eyebrow='UID', judul='Transfer EXIM',
        icon='arrows-exchange', desc='Monitoring ekspor dan impor energi antar unit.')

@app.route('/rekap')
def halaman_rekap():
    return render_template('rekap.html',
        eyebrow='Rekap kWh', judul='Rekap kWh',
        icon='report', desc='Rekap kWh dan susut per GI per bulan.')


@app.route('/kwh-jual')
def halaman_kwh_jual():
    return render_template('kwh_jual.html',
        eyebrow='Transaksi', judul='kWh Jual',
        icon='receipt-2', desc='Data transaksi kWh jual pelanggan sebagai referensi alokasi dan transfer.')


@app.route('/emin')
def halaman_emin():
    return render_template('rekap.html',
        eyebrow='Transaksi', judul='EMIN',
        icon='file-analytics', desc='Data EMIN sebagai pendukung transaksi dan rekonsiliasi energi.')


@app.route('/profile')
def halaman_profile():
    return render_template('profile.html',
        eyebrow='Akun', judul='Profile',
        icon='user-circle', desc='Informasi akun dan preferensi akses aplikasi.')


@app.route('/master-data')
@role_required('admin', 'operator')
def halaman_master_data():
    return render_template('master_data.html',
        eyebrow='Data Master', judul='Master Data',
        icon='database', desc='Kelola master GI, trafo, penyulang, dan area/UP3.')


@app.route('/upload')
@role_required('admin', 'operator')
def halaman_upload():
    return render_template('upload.html',
        eyebrow='Laporan', judul='Upload NKWh',
        icon='upload', desc='Upload file NKWh Excel/CSV untuk import data penyulang.')


@app.route('/security')
@role_required('admin')
def halaman_security():
    return render_template('security.html',
        eyebrow='Admin', judul='Security',
        icon='shield-lock', desc='Manajemen user dan audit log aplikasi.')


# ════════════════════════════════════════════════
# API — SIDEBAR STATS (diload di base.html)
# ════════════════════════════════════════════════

@app.route('/api/sidebar-stats')
def api_sidebar_stats():
    """
    Data mini stats sidebar: jumlah GI aktif dan jumlah alert.
    Ringan — tidak join banyak tabel.
    """
    try:
        gi_aktif = GarduInduk.query.filter_by(aktif=True).count()

        # Alert = FeederReading dengan flag_alert=True bulan ini
        from datetime import date
        bulan_ini = date.today().replace(day=1)
        alert_count = FeederReading.query.filter_by(
            flag_alert=True,
            periode_bulan=bulan_ini
        ).count()

        return jsonify({
            'gi_aktif':    gi_aktif,
            'alert_count': alert_count
        })
    except Exception as e:
        return jsonify({'gi_aktif': 0, 'alert_count': 0, 'error': str(e)})


# ════════════════════════════════════════════════
# API — MASTER DATA
# ════════════════════════════════════════════════

def _master_writer_required():
    user = getattr(g, 'current_user', None)
    if not user or user.role not in WRITE_ROLES:
        return _json_error('Akses ubah master data hanya untuk admin/operator.', 403)
    return None


@app.route('/api/monthly-status')
@role_required('admin', 'operator')
def api_monthly_status_list():
    try:
        tahun = request.args.get('tahun', default=date.today().year, type=int)
        start = date(tahun, 1, 1)
        end = date(tahun + 1, 1, 1)
        rows = MonthlyDataStatus.query.filter(
            MonthlyDataStatus.periode_bulan >= start,
            MonthlyDataStatus.periode_bulan < end,
        ).all()
        by_month = {row.periode_bulan.month: row for row in rows}
        payload_rows = [
            _workflow_payload(date(tahun, month, 1), by_month.get(month))
            for month in range(1, 13)
        ]
        return jsonify({'tahun': tahun, 'rows': payload_rows})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/monthly-status/<periode>', methods=['GET', 'PATCH', 'POST'])
@role_required('admin', 'operator')
def api_monthly_status_detail(periode):
    try:
        period = _workflow_period(periode)
        record = _workflow_record(period, create=request.method != 'GET')

        if request.method == 'GET':
            return jsonify(_workflow_payload(period, record))

        denied = _master_writer_required()
        if denied:
            return denied

        payload = _request_payload()
        new_status = _normalize_workflow_status(payload.get('status'))
        current_status = _normalize_workflow_status(record.status)
        user = getattr(g, 'current_user', None)
        allowed = _workflow_allowed_statuses(current_status, user)
        if new_status not in allowed:
            return _json_error(
                f'Transisi status dari {WORKFLOW_STATUS_LABELS[current_status]} '
                f'ke {WORKFLOW_STATUS_LABELS[new_status]} tidak diizinkan.',
                400,
            )
        if new_status == 'TERKUNCI' and (not user or user.role != 'admin'):
            return _json_error('Hanya admin yang bisa mengunci periode.', 403)

        force_finalize = _bool_value(payload.get('force_finalize') or payload.get('force'))
        readiness = None
        if new_status in {'FINAL', 'TERKUNCI'}:
            readiness = _readiness_payload(period)
            if not readiness['can_finalize']:
                can_override = user and user.role == 'admin' and force_finalize
                note = _clean_value(payload.get('catatan'))
                if not can_override:
                    return jsonify({
                        'error': 'Data wajib periode ini belum lengkap untuk Final/Terkunci.',
                        'blockers': readiness['blockers'],
                        'readiness': readiness,
                    }), 409
                if not note:
                    return jsonify({
                        'error': 'Catatan wajib diisi untuk override Final/Terkunci.',
                        'blockers': readiness['blockers'],
                        'readiness': readiness,
                    }), 400

        record.status = new_status
        record.catatan = _clean_value(payload.get('catatan')) or None
        if new_status == 'TERKUNCI':
            record.locked_at = datetime.utcnow()
            record.locked_by = user.username if user else None
        else:
            record.locked_at = None
            record.locked_by = None

        _audit('UPDATE_MONTHLY_STATUS', entity_type='monthly_data_status', entity_id=record.id, detail={
            'periode_bulan': period.strftime('%Y-%m-%d'),
            'from_status': current_status,
            'to_status': new_status,
            'force_finalize': force_finalize,
            'readiness_score': readiness['score'] if readiness else None,
            'readiness_blockers': readiness['blockers'] if readiness else [],
        })
        db.session.commit()
        return jsonify(_workflow_payload(period, record))
    except ValueError as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@app.route('/api/monthly-status/<periode>/activity')
@role_required('admin', 'operator')
def api_monthly_status_activity(periode):
    try:
        period = _workflow_period(periode)
        limit = min(request.args.get('limit', default=30, type=int), 100)
        return jsonify(_monthly_activity_payload(period, limit))
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/monthly-status/<periode>/readiness')
@role_required('admin', 'operator')
def api_monthly_status_readiness(periode):
    try:
        period = _workflow_period(periode)
        return jsonify(_readiness_payload(period))
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/monthly-status/<periode>/audit-package')
@role_required('admin', 'operator', 'auditor')
def api_monthly_status_audit_package(periode):
    try:
        period = _workflow_period(periode)
        workflow = _workflow_payload(period)
        readiness = _readiness_payload(period)
        activity = _monthly_activity_payload(period, 100)
        user = getattr(g, 'current_user', None)
        return jsonify({
            'generated_at': datetime.utcnow().isoformat() + 'Z',
            'generated_by': user.username if user else None,
            'periode': period.strftime('%Y-%m'),
            'periode_bulan': period.strftime('%Y-%m-%d'),
            'summary': {
                'workflow_status': workflow['status'],
                'workflow_label': workflow['label'],
                'readiness_score': readiness['score'],
                'can_finalize': readiness['can_finalize'],
                'blocker_count': len(readiness['blockers']),
                'alert_count': readiness['alert_count'],
                'activity_count': len(activity['rows']),
            },
            'workflow': workflow,
            'readiness': readiness,
            'activity': activity['rows'],
        })
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/export/<module>.<fmt>')
@role_required('admin', 'operator', 'auditor')
def api_export_report(module, fmt):
    try:
        return _report_file_response(module, fmt)
    except ValueError as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


def _clean_value(value, default=''):
    text_value = str(value or '').strip()
    return text_value if text_value else default


def _decimal_payload(value, default='0'):
    if value in (None, ''):
        return Decimal(default)
    return Decimal(str(value))


@app.route('/api/master-data/summary')
def api_master_summary():
    try:
        active_gi = GarduInduk.query.filter_by(aktif=True).count()
        active_trafo = Trafo.query.filter_by(aktif=True).count()
        active_penyulang = Penyulang.query.filter_by(aktif=True).count()
        active_area = AreaUnit.query.filter_by(aktif=True).count()
        missing_area = Penyulang.query.filter(Penyulang.aktif.is_(True)).filter(
            (Penyulang.area_up3.is_(None)) | (Penyulang.area_up3 == '')
        ).count()
        trafo_without_feeder = sum(1 for trafo in Trafo.query.filter_by(aktif=True).all() if not trafo.penyulangs)
        return jsonify({
            'gi': active_gi,
            'trafo': active_trafo,
            'penyulang': active_penyulang,
            'area_unit': active_area,
            'missing_area': missing_area,
            'trafo_without_feeder': trafo_without_feeder,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/area-unit', methods=['GET', 'POST'])
def api_area_unit():
    try:
        if request.method == 'POST':
            denied = _master_writer_required()
            if denied:
                return denied
            payload = _request_payload()
            kode = _clean_value(payload.get('kode_unit')).upper()
            nama = _clean_value(payload.get('nama_unit'))
            if not kode or not nama:
                return _json_error('Kode unit dan nama unit wajib diisi.', 400)
            if AreaUnit.query.filter_by(kode_unit=kode).first():
                return _json_error('Kode unit sudah terdaftar.', 409)
            unit = AreaUnit(
                kode_unit=kode,
                nama_unit=nama,
                jenis=_clean_value(payload.get('jenis'), 'UP3').upper(),
                parent_unit=_clean_value(payload.get('parent_unit')) or None,
                aktif=_bool_value(payload.get('aktif', True)),
            )
            db.session.add(unit)
            _audit('CREATE_AREA_UNIT', entity_type='area_unit', detail={'kode_unit': kode})
            db.session.commit()
            return jsonify(unit.to_dict()), 201

        q = AreaUnit.query
        if request.args.get('all') != '1':
            q = q.filter_by(aktif=True)
        return jsonify([u.to_dict() for u in q.order_by(AreaUnit.jenis, AreaUnit.nama_unit).all()])
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@app.route('/api/area-unit/<int:unit_id>', methods=['PATCH', 'POST'])
def api_area_unit_update(unit_id):
    denied = _master_writer_required()
    if denied:
        return denied
    unit = db.session.get(AreaUnit, unit_id)
    if not unit:
        return _json_error('Area/unit tidak ditemukan.', 404)
    payload = _request_payload()
    kode = _clean_value(payload.get('kode_unit'), unit.kode_unit).upper()
    nama = _clean_value(payload.get('nama_unit'), unit.nama_unit)
    existing = AreaUnit.query.filter(AreaUnit.kode_unit == kode, AreaUnit.id != unit.id).first()
    if existing:
        return _json_error('Kode unit sudah dipakai area/unit lain.', 409)
    before = unit.to_dict()
    unit.kode_unit = kode
    unit.nama_unit = nama
    unit.jenis = _clean_value(payload.get('jenis'), unit.jenis).upper()
    unit.parent_unit = _clean_value(payload.get('parent_unit')) or None
    unit.aktif = _bool_value(payload.get('aktif', unit.aktif))
    _audit('UPDATE_AREA_UNIT', entity_type='area_unit', entity_id=unit.id, detail={
        'before': before,
        'after': unit.to_dict(),
    })
    db.session.commit()
    return jsonify(unit.to_dict())


@app.route('/api/gardu-induk', methods=['GET', 'POST'])
def api_gardu_induk():
    try:
        if request.method == 'POST':
            denied = _master_writer_required()
            if denied:
                return denied
            payload = _request_payload()
            kode = _clean_value(payload.get('kode_gi')).upper()
            nama = _clean_value(payload.get('nama_gi'))
            if not kode or not nama:
                return _json_error('Kode GI dan nama GI wajib diisi.', 400)
            if GarduInduk.query.filter_by(kode_gi=kode).first():
                return _json_error('Kode GI sudah terdaftar.', 409)
            gi = GarduInduk(
                kode_gi=kode,
                nama_gi=nama,
                area=_clean_value(payload.get('area')) or None,
                unit=_clean_value(payload.get('unit')) or None,
                alamat=_clean_value(payload.get('alamat')) or None,
                aktif=_bool_value(payload.get('aktif', True)),
            )
            db.session.add(gi)
            _audit('CREATE_GI', entity_type='gardu_induk', detail={'kode_gi': kode})
            db.session.commit()
            return jsonify(gi.to_dict()), 201

        q = GarduInduk.query
        if request.args.get('all') != '1':
            q = q.filter_by(aktif=True)
        gis = q.order_by(GarduInduk.nama_gi).all()
        return jsonify([g.to_dict() for g in gis])
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@app.route('/api/gardu-induk/<int:gi_id>', methods=['PATCH', 'POST'])
def api_gardu_induk_update(gi_id):
    denied = _master_writer_required()
    if denied:
        return denied
    gi = db.session.get(GarduInduk, gi_id)
    if not gi:
        return _json_error('Gardu induk tidak ditemukan.', 404)
    payload = _request_payload()
    kode = _clean_value(payload.get('kode_gi'), gi.kode_gi).upper()
    nama = _clean_value(payload.get('nama_gi'), gi.nama_gi)
    existing = GarduInduk.query.filter(GarduInduk.kode_gi == kode, GarduInduk.id != gi.id).first()
    if existing:
        return _json_error('Kode GI sudah dipakai gardu induk lain.', 409)
    before = gi.to_dict()
    gi.kode_gi = kode
    gi.nama_gi = nama
    gi.area = _clean_value(payload.get('area')) or None
    gi.unit = _clean_value(payload.get('unit')) or None
    gi.alamat = _clean_value(payload.get('alamat')) or None
    gi.aktif = _bool_value(payload.get('aktif', gi.aktif))
    _audit('UPDATE_GI', entity_type='gardu_induk', entity_id=gi.id, detail={
        'before': before,
        'after': gi.to_dict(),
    })
    db.session.commit()
    return jsonify(gi.to_dict())


@app.route('/api/trafo', methods=['GET', 'POST'])
def api_trafo():
    try:
        if request.method == 'POST':
            denied = _master_writer_required()
            if denied:
                return denied
            payload = _request_payload()
            gi = db.session.get(GarduInduk, int(payload.get('gi_id') or 0))
            if not gi:
                return _json_error('Gardu induk wajib dipilih.', 400)
            kode = _clean_value(payload.get('kode_trafo')).upper()
            nama = _clean_value(payload.get('nama_trafo'))
            if not kode or not nama:
                return _json_error('Kode trafo dan nama trafo wajib diisi.', 400)
            if Trafo.query.filter_by(gi_id=gi.id, kode_trafo=kode).first():
                return _json_error('Kode trafo sudah ada di GI ini.', 409)
            trafo = Trafo(
                gi_id=gi.id,
                kode_trafo=kode,
                nama_trafo=nama,
                kapasitas_mva=_decimal_payload(payload.get('kapasitas_mva'), '0'),
                tegangan_kv=_decimal_payload(payload.get('tegangan_kv'), '20'),
                aktif=_bool_value(payload.get('aktif', True)),
            )
            db.session.add(trafo)
            _audit('CREATE_TRAFO', entity_type='trafo', detail={'kode_trafo': kode, 'gi_id': gi.id})
            db.session.commit()
            return jsonify(trafo.to_dict()), 201

        gi_id = request.args.get('gi_id', type=int)
        q = Trafo.query
        if request.args.get('all') != '1':
            q = q.filter_by(aktif=True)
        if gi_id:
            q = q.filter_by(gi_id=gi_id)
        return jsonify([t.to_dict() for t in q.order_by(Trafo.kode_trafo).all()])
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@app.route('/api/trafo/<int:trafo_id>', methods=['PATCH', 'POST'])
def api_trafo_update(trafo_id):
    denied = _master_writer_required()
    if denied:
        return denied
    trafo = db.session.get(Trafo, trafo_id)
    if not trafo:
        return _json_error('Trafo tidak ditemukan.', 404)
    payload = _request_payload()
    gi = db.session.get(GarduInduk, int(payload.get('gi_id') or trafo.gi_id))
    if not gi:
        return _json_error('Gardu induk tidak ditemukan.', 400)
    kode = _clean_value(payload.get('kode_trafo'), trafo.kode_trafo).upper()
    nama = _clean_value(payload.get('nama_trafo'), trafo.nama_trafo)
    exists = Trafo.query.filter(Trafo.gi_id == gi.id, Trafo.kode_trafo == kode, Trafo.id != trafo.id).first()
    if exists:
        return _json_error('Kode trafo sudah ada di GI ini.', 409)
    before = trafo.to_dict()
    trafo.gi_id = gi.id
    trafo.kode_trafo = kode
    trafo.nama_trafo = nama
    trafo.kapasitas_mva = _decimal_payload(payload.get('kapasitas_mva'), str(trafo.kapasitas_mva or 0))
    trafo.tegangan_kv = _decimal_payload(payload.get('tegangan_kv'), str(trafo.tegangan_kv or 20))
    trafo.aktif = _bool_value(payload.get('aktif', trafo.aktif))
    db.session.flush()
    _audit('UPDATE_TRAFO', entity_type='trafo', entity_id=trafo.id, detail={
        'before': before,
        'after': trafo.to_dict(),
    })
    db.session.commit()
    return jsonify(trafo.to_dict())


@app.route('/api/penyulang', methods=['GET', 'POST'])
def api_penyulang_list():
    try:
        if request.method == 'POST':
            denied = _master_writer_required()
            if denied:
                return denied
            payload = _request_payload()
            trafo = db.session.get(Trafo, int(payload.get('trafo_id') or 0))
            if not trafo:
                return _json_error('Trafo wajib dipilih.', 400)
            kode = _clean_value(payload.get('kode_penyulang')).upper()
            nama = _clean_value(payload.get('nama_penyulang'))
            if not kode or not nama:
                return _json_error('Kode penyulang dan nama penyulang wajib diisi.', 400)
            if Penyulang.query.filter_by(trafo_id=trafo.id, kode_penyulang=kode).first():
                return _json_error('Kode penyulang sudah ada di trafo ini.', 409)
            status_value = _clean_value(payload.get('status'), 'AKTIF').upper()
            penyulang = Penyulang(
                trafo_id=trafo.id,
                gi_id=trafo.gi_id,
                kode_penyulang=kode,
                nama_penyulang=nama,
                jenis=_clean_value(payload.get('jenis'), 'REGULAR').upper(),
                area_up3=_clean_value(payload.get('area_up3')) or None,
                ex_cabang=_clean_value(payload.get('ex_cabang')) or None,
                status=status_value,
                aktif=_bool_value(payload.get('aktif', status_value != 'NONAKTIF')),
            )
            db.session.add(penyulang)
            _audit('CREATE_PENYULANG', entity_type='penyulang', detail={'kode_penyulang': kode, 'trafo_id': trafo.id})
            db.session.commit()
            return jsonify(penyulang.to_dict()), 201

        trafo_id = request.args.get('trafo_id', type=int)
        gi_id    = request.args.get('gi_id',    type=int)
        area_up3 = request.args.get('area_up3', '').strip()
        status   = request.args.get('status', '').strip()
        q = Penyulang.query
        if not status and request.args.get('all') != '1':
            q = q.filter_by(aktif=True)
        if trafo_id:
            q = q.filter_by(trafo_id=trafo_id)
        if gi_id:
            q = q.filter_by(gi_id=gi_id)
        if area_up3:
            q = q.filter(Penyulang.area_up3 == area_up3)
        if status:
            q = q.filter(Penyulang.status == status)
        return jsonify([p.to_dict() for p in q.order_by(Penyulang.kode_penyulang).all()])
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@app.route('/api/penyulang/<int:penyulang_id>', methods=['PATCH', 'POST'])
def api_penyulang_update(penyulang_id):
    denied = _master_writer_required()
    if denied:
        return denied
    penyulang = db.session.get(Penyulang, penyulang_id)
    if not penyulang:
        return _json_error('Penyulang tidak ditemukan.', 404)
    payload = _request_payload()
    trafo = db.session.get(Trafo, int(payload.get('trafo_id') or penyulang.trafo_id))
    if not trafo:
        return _json_error('Trafo tidak ditemukan.', 400)
    kode = _clean_value(payload.get('kode_penyulang'), penyulang.kode_penyulang).upper()
    nama = _clean_value(payload.get('nama_penyulang'), penyulang.nama_penyulang)
    exists = Penyulang.query.filter(
        Penyulang.trafo_id == trafo.id,
        Penyulang.kode_penyulang == kode,
        Penyulang.id != penyulang.id,
    ).first()
    if exists:
        return _json_error('Kode penyulang sudah ada di trafo ini.', 409)
    before = penyulang.to_dict()
    status_value = _clean_value(payload.get('status'), penyulang.status or 'AKTIF').upper()
    penyulang.trafo_id = trafo.id
    penyulang.gi_id = trafo.gi_id
    penyulang.kode_penyulang = kode
    penyulang.nama_penyulang = nama
    penyulang.jenis = _clean_value(payload.get('jenis'), penyulang.jenis or 'REGULAR').upper()
    penyulang.area_up3 = _clean_value(payload.get('area_up3')) or None
    penyulang.ex_cabang = _clean_value(payload.get('ex_cabang')) or None
    penyulang.status = status_value
    penyulang.aktif = _bool_value(payload.get('aktif', penyulang.aktif))
    db.session.flush()
    _audit('UPDATE_PENYULANG', entity_type='penyulang', entity_id=penyulang.id, detail={
        'before': before,
        'after': penyulang.to_dict(),
    })
    db.session.commit()
    return jsonify(penyulang.to_dict())


@app.route('/api/penyulang-area')
def api_penyulang_area():
    try:
        rows = db.session.query(Penyulang.area_up3)\
            .filter(Penyulang.aktif.is_(True))\
            .filter(Penyulang.area_up3.isnot(None))\
            .filter(Penyulang.area_up3 != '')\
            .distinct()\
            .order_by(Penyulang.area_up3)\
            .all()
        return jsonify([r[0] for r in rows])
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ════════════════════════════════════════════════
# API — DASHBOARD
# ════════════════════════════════════════════════

@app.route('/api/dashboard-data')
def get_dashboard_data():
    try:
        tahun = request.args.get('tahun', type=int)

        sub_mu = db.session.query(
            MeterReading.periode_bulan,
            func.sum(
                func.coalesce(MeterReading.mu_kwh_wbp,   0) +
                func.coalesce(MeterReading.mu_kwh_lwbp1, 0) +
                func.coalesce(MeterReading.mu_kwh_lwbp2, 0)
            ).label('total_mu')
        ).group_by(MeterReading.periode_bulan)

        sub_py = db.session.query(
            FeederReading.periode_bulan,
            func.sum(
                func.coalesce(FeederReading.kwh_wbp,   0) +
                func.coalesce(FeederReading.kwh_lwbp1, 0) +
                func.coalesce(FeederReading.kwh_lwbp2, 0)
            ).label('total_penyulang')
        ).group_by(FeederReading.periode_bulan)

        if tahun:
            sub_mu = sub_mu.filter(func.extract('year', MeterReading.periode_bulan) == tahun)
            sub_py = sub_py.filter(func.extract('year', FeederReading.periode_bulan) == tahun)

        sub_mu = sub_mu.subquery()
        sub_py = sub_py.subquery()

        rows = db.session.query(
            sub_mu.c.periode_bulan,
            sub_mu.c.total_mu,
            func.coalesce(sub_py.c.total_penyulang, 0).label('total_penyulang')
        ).outerjoin(
            sub_py, sub_mu.c.periode_bulan == sub_py.c.periode_bulan
        ).order_by(sub_mu.c.periode_bulan).all()

        data = []
        t_mu = t_py = 0
        for r in rows:
            mu  = float(r.total_mu)
            py  = float(r.total_penyulang)
            sk  = mu - py
            pct = round(sk / mu * 100, 2) if mu > 0 else 0
            t_mu += mu; t_py += py
            data.append({
                'tanggal':          r.periode_bulan.strftime('%Y-%m-%d'),
                'meter_utama':      mu,
                'total_penyulang':  py,
                'susut_kwh':        round(sk, 2),
                'persentase_susut': pct,
            })

        t_sk  = t_mu - t_py
        t_pct = round(t_sk / t_mu * 100, 2) if t_mu > 0 else 0
        return jsonify({
            'data_bulanan': data,
            'total': {
                'meter_utama':      t_mu,
                'total_penyulang':  t_py,
                'total_susut':      round(t_sk, 2),
                'persentase_total': t_pct,
            }
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ════════════════════════════════════════════════
# API — FEEDER, METER, TRANSFER, REKAP
# ════════════════════════════════════════════════

@app.route('/api/executive-dashboard')
def api_executive_dashboard():
    try:
        tahun = request.args.get('tahun', type=int) or date.today().year
        month = request.args.get('month', type=int) or date.today().month
        if month < 1 or month > 12:
            return jsonify({'error': 'Bulan tidak valid.'}), 400
        period = date(tahun, month, 1)

        mu_expr = _kwh_sum(MeterReading, MeterReading.mu_kwh_wbp, MeterReading.mu_kwh_lwbp1, MeterReading.mu_kwh_lwbp2)
        feeder_expr = _kwh_sum(FeederReading, FeederReading.kwh_wbp, FeederReading.kwh_lwbp1, FeederReading.kwh_lwbp2)

        total_masuk = _float_value(
            db.session.query(func.sum(mu_expr))
            .filter(MeterReading.periode_bulan == period)
            .scalar()
        )
        total_keluar = _float_value(
            db.session.query(func.sum(feeder_expr))
            .filter(FeederReading.periode_bulan == period)
            .scalar()
        )
        susut = total_masuk - total_keluar
        susut_pct = (susut / total_masuk * 100) if total_masuk else 0

        mu_rows = db.session.query(
            MeterReading.gi_id,
            func.sum(mu_expr).label('total_mu'),
        ).filter(
            MeterReading.periode_bulan == period,
        ).group_by(MeterReading.gi_id).all()
        feeder_rows = db.session.query(
            FeederReading.gi_id,
            func.sum(feeder_expr).label('total_feeder'),
        ).filter(
            FeederReading.periode_bulan == period,
        ).group_by(FeederReading.gi_id).all()
        gi_names = {
            gi.id: gi.nama_gi
            for gi in GarduInduk.query.filter_by(aktif=True).all()
        }
        mu_by_gi = {row.gi_id: _float_value(row.total_mu) for row in mu_rows}
        feeder_by_gi = {row.gi_id: _float_value(row.total_feeder) for row in feeder_rows}
        gi_deviasi = []
        for gi_id in sorted(set(mu_by_gi) | set(feeder_by_gi)):
            mu = mu_by_gi.get(gi_id, 0)
            feeder = feeder_by_gi.get(gi_id, 0)
            gap = mu - feeder
            gi_deviasi.append({
                'gi_id': gi_id,
                'nama_gi': gi_names.get(gi_id, f'GI #{gi_id}'),
                'meter_utama': round(mu, 2),
                'penyulang': round(feeder, 2),
                'deviasi_kwh': round(gap, 2),
                'deviasi_persen': round((gap / mu * 100) if mu else 0, 2),
            })
        gi_deviasi.sort(key=lambda row: abs(row['deviasi_persen']), reverse=True)

        anomaly_rows = db.session.query(
            FeederReading, Penyulang, Trafo, GarduInduk
        ).join(
            Penyulang, FeederReading.penyulang_id == Penyulang.id
        ).join(
            Trafo, FeederReading.trafo_id == Trafo.id
        ).join(
            GarduInduk, FeederReading.gi_id == GarduInduk.id
        ).filter(
            FeederReading.periode_bulan == period
        ).all()
        anomalies = []
        for reading, penyulang, trafo, gi in anomaly_rows:
            pct = _float_value(reading.deviasi_persen)
            if not reading.flag_alert and abs(pct) < 20:
                continue
            anomalies.append({
                'penyulang': penyulang.nama_penyulang,
                'kode_penyulang': penyulang.kode_penyulang,
                'gardu_induk': gi.nama_gi,
                'trafo': trafo.kode_trafo,
                'area_up3': penyulang.area_up3 or 'Belum Dipetakan',
                'kwh_total': round(reading.kwh_total, 2),
                'deviasi_persen': round(pct, 2),
                'anomaly_type': reading.anomaly_type or ('Naik/Turun Tidak Wajar' if reading.flag_alert else 'Deviasi Tinggi'),
            })
        anomalies.sort(key=lambda row: abs(row['deviasi_persen']), reverse=True)

        readiness = _readiness_payload(period)
        workflow = _workflow_payload(period)
        return jsonify({
            'periode': period.strftime('%Y-%m'),
            'periode_bulan': period.strftime('%Y-%m-%d'),
            'total_kwh_masuk': round(total_masuk, 2),
            'total_kwh_keluar': round(total_keluar, 2),
            'susut_kwh': round(susut, 2),
            'susut_persen': round(susut_pct, 2),
            'gi_deviasi_terbesar': gi_deviasi[:5],
            'penyulang_anomali': anomalies[:8],
            'readiness': readiness,
            'workflow': workflow,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/feeder-data')
def api_feeder_data():
    try:
        gi_id    = request.args.get('gi_id',    type=int)
        trafo_id = request.args.get('trafo_id', type=int)
        bulan    = request.args.get('bulan', '').strip()
        
        q = db.session.query(FeederReading, Penyulang)\
              .join(Penyulang, FeederReading.penyulang_id == Penyulang.id)
        if gi_id:    q = q.filter(FeederReading.gi_id    == gi_id)
        if trafo_id: q = q.filter(FeederReading.trafo_id == trafo_id)
        if bulan:
            try:
                thn, bln = bulan.split('-')
                thn, bln = int(thn), int(bln)
                if not (1 <= bln <= 12) or thn < 2000 or thn > 2100:
                    return jsonify({'error': 'Invalid date format'}), 400
                q = q.filter(
                    func.extract('year',  FeederReading.periode_bulan) == thn,
                    func.extract('month', FeederReading.periode_bulan) == bln
                )
            except (ValueError, AttributeError):
                return jsonify({'error': 'Format bulan harus YYYY-MM'}), 400
        
        result = []
        for fr, py in q.order_by(Penyulang.kode_penyulang).all():
            d = fr.to_dict()
            d['kode_penyulang'] = py.kode_penyulang
            d['nama_penyulang'] = py.nama_penyulang
            d['jenis'] = py.jenis
            d['area_up3'] = py.area_up3
            d['ex_cabang'] = py.ex_cabang
            d['status'] = py.status or ('AKTIF' if py.aktif else 'NONAKTIF')
            result.append(d)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/meter-data')
def api_meter_data():
    try:
        gi_id    = request.args.get('gi_id',    type=int)
        trafo_id = request.args.get('trafo_id', type=int)
        bulan    = request.args.get('bulan')

        q = db.session.query(MeterReading, Trafo, GarduInduk)\
              .join(Trafo,      MeterReading.trafo_id == Trafo.id)\
              .join(GarduInduk, MeterReading.gi_id    == GarduInduk.id)
        if gi_id:    q = q.filter(MeterReading.gi_id    == gi_id)
        if trafo_id: q = q.filter(MeterReading.trafo_id == trafo_id)
        if bulan:
            thn, bln = bulan.split('-')
            q = q.filter(
                func.extract('year',  MeterReading.periode_bulan) == int(thn),
                func.extract('month', MeterReading.periode_bulan) == int(bln)
            )

        result = []
        for mr, tr, gi in q.order_by(MeterReading.periode_bulan).all():
            d = mr.to_dict()
            d['kode_trafo'] = tr.kode_trafo
            d['nama_trafo'] = tr.nama_trafo
            d['nama_gi']    = gi.nama_gi
            result.append(d)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/transfer-data')
def api_transfer_data():
    try:
        bulan = request.args.get('bulan')
        tahun = request.args.get('tahun', type=int)
        q = TransferAntarUnit.query
        if bulan:
            thn, bln = bulan.split('-')
            q = q.filter(
                func.extract('year',  TransferAntarUnit.periode_bulan) == int(thn),
                func.extract('month', TransferAntarUnit.periode_bulan) == int(bln)
            )
        elif tahun:
            q = q.filter(func.extract('year', TransferAntarUnit.periode_bulan) == tahun)
        return jsonify([r.to_dict() for r in q.order_by(TransferAntarUnit.periode_bulan).all()])
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/rekap')
def api_rekap():
    try:
        tahun = request.args.get('tahun', type=int)
        gi_id = request.args.get('gi_id', type=int)
        q = db.session.query(RekapBulanan, GarduInduk)\
              .join(GarduInduk, RekapBulanan.gi_id == GarduInduk.id)\
              .filter(RekapBulanan.trafo_id.is_(None))
        if tahun: q = q.filter(func.extract('year', RekapBulanan.periode_bulan) == tahun)
        if gi_id: q = q.filter(RekapBulanan.gi_id == gi_id)
        result = []
        for rb, gi in q.order_by(RekapBulanan.periode_bulan).all():
            d = rb.to_dict()
            d['nama_gi'] = gi.nama_gi
            d['kode_gi'] = gi.kode_gi
            result.append(d)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ════════════════════════════════════════════════
# API — UPLOAD (placeholder, lengkap di Sesi 2)
# ════════════════════════════════════════════════

@app.route('/api/audit-log')
@role_required('admin')
def api_audit_log():
    try:
        limit = request.args.get('limit', default=100, type=int)
        limit = min(max(limit, 1), 500)
        rows = AuditLog.query.order_by(AuditLog.created_at.desc()).limit(limit).all()
        return jsonify([row.to_dict() for row in rows])
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/security-summary')
@role_required('admin')
def api_security_summary():
    try:
        users_total = User.query.count()
        active_users = User.query.filter_by(aktif=True).count()
        failed_logins = AuditLog.query.filter_by(action='LOGIN_FAILED').count()
        imports = AuditLog.query.filter(
            AuditLog.action.in_(['IMPORT_NKWH', 'IMPORT_PENYULANG'])
        ).count()
        return jsonify({
            'users_total': users_total,
            'active_users': active_users,
            'failed_logins': failed_logins,
            'imports': imports,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/module-access')
@role_required('admin')
def api_module_access():
    try:
        return jsonify(_module_access_payload(request.args.get('role')))
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/users')
@role_required('admin')
def api_users_list():
    try:
        rows = User.query.order_by(User.username).all()
        return jsonify([row.to_dict() for row in rows])
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/users', methods=['POST'])
@role_required('admin')
def api_users_create():
    payload = _request_payload()
    username = (payload.get('username') or '').strip()
    password = payload.get('password') or ''
    role = (payload.get('role') or 'viewer').strip().lower()
    if not username:
        return jsonify({'error': 'Username wajib diisi.'}), 400
    if role not in ROLES:
        return jsonify({'error': 'Role tidak valid.'}), 400
    password_error = _validate_password_policy(password)
    if password_error:
        return jsonify({'error': password_error}), 400
    if User.query.filter_by(username=username).first():
        return jsonify({'error': 'Username sudah digunakan.'}), 409

    user = User(
        username=username,
        nama_lengkap=(payload.get('nama_lengkap') or '').strip() or username,
        email=(payload.get('email') or '').strip() or None,
        role=role,
        aktif=_bool_value(payload.get('aktif', True)),
    )
    user.set_password(password)
    db.session.add(user)
    db.session.flush()
    _audit('CREATE_USER', entity_type='user', entity_id=user.id, detail=user.to_dict())
    db.session.commit()
    return jsonify(user.to_dict()), 201


@app.route('/api/users/<int:user_id>', methods=['PATCH', 'POST'])
@role_required('admin')
def api_users_update(user_id):
    user = db.session.get(User, user_id)
    if not user:
        return jsonify({'error': 'User tidak ditemukan.'}), 404
    payload = _request_payload()
    role = (payload.get('role') or user.role).strip().lower()
    if role not in ROLES:
        return jsonify({'error': 'Role tidak valid.'}), 400
    admin_count = User.query.filter_by(role='admin', aktif=True).count()
    new_active = _bool_value(payload.get('aktif', user.aktif))
    if user.role == 'admin' and (role != 'admin' or not new_active) and admin_count <= 1:
        return jsonify({'error': 'Minimal harus ada satu admin aktif.'}), 400

    before = user.to_dict()
    user.nama_lengkap = (payload.get('nama_lengkap') or user.nama_lengkap or user.username).strip()
    user.email = (payload.get('email') or '').strip() or None
    user.role = role
    user.aktif = new_active
    db.session.flush()
    _audit('UPDATE_USER', entity_type='user', entity_id=user.id, detail={
        'before': before,
        'after': user.to_dict(),
    })
    db.session.commit()
    return jsonify(user.to_dict())


@app.route('/api/users/<int:user_id>/password', methods=['POST'])
@role_required('admin')
def api_users_reset_password(user_id):
    user = db.session.get(User, user_id)
    if not user:
        return jsonify({'error': 'User tidak ditemukan.'}), 404
    payload = _request_payload()
    password = payload.get('password') or ''
    password_error = _validate_password_policy(password)
    if password_error:
        return jsonify({'error': password_error}), 400
    user.set_password(password)
    db.session.flush()
    _audit('RESET_USER_PASSWORD', entity_type='user', entity_id=user.id, detail={'username': user.username})
    db.session.commit()
    return jsonify({'message': 'Password berhasil direset.'})


@app.route('/api/me', methods=['GET', 'PATCH', 'POST'])
def api_me_profile():
    user = getattr(g, 'current_user', None)
    if not user:
        return jsonify({'error': 'Login diperlukan.'}), 401
    if request.method == 'GET':
        return jsonify(user.to_dict())
    payload = _request_payload()
    before = user.to_dict()
    user.nama_lengkap = (payload.get('nama_lengkap') or user.nama_lengkap or user.username).strip()
    user.email = (payload.get('email') or '').strip() or None
    db.session.flush()
    _audit('UPDATE_OWN_PROFILE', entity_type='user', entity_id=user.id, detail={
        'before': before,
        'after': user.to_dict(),
    })
    db.session.commit()
    return jsonify(user.to_dict())


@app.route('/api/me/password', methods=['POST'])
def api_change_own_password():
    user = getattr(g, 'current_user', None)
    if not user:
        return jsonify({'error': 'Login diperlukan.'}), 401
    payload = _request_payload()
    current_password = payload.get('current_password') or ''
    new_password = payload.get('new_password') or ''
    if not user.check_password(current_password):
        return jsonify({'error': 'Password lama tidak sesuai.'}), 400
    password_error = _validate_password_policy(new_password)
    if password_error:
        return jsonify({'error': password_error}), 400
    user.set_password(new_password)
    db.session.flush()
    _audit('CHANGE_OWN_PASSWORD', entity_type='user', entity_id=user.id, detail={'username': user.username})
    db.session.commit()
    return jsonify({'message': 'Password berhasil diganti.'})


def _shift_month(period, offset):
    month_index = period.year * 12 + period.month - 1 + offset
    return date(month_index // 12, month_index % 12 + 1, 1)


def _kwh_jual_catalog_payload():
    return {
        'groups': KWH_JUAL_GROUP_LABELS,
        'rows': KWH_JUAL_CATALOG,
    }


def _kwh_jual_payload(gi_id, period):
    q = KwhJual.query.filter(KwhJual.periode_bulan == period)
    if gi_id:
        q = q.filter(KwhJual.gi_id == gi_id)

    values_by_sub = defaultdict(float)
    for row in q.all():
        if row.sub_golongan in KWH_JUAL_SUB_INDEX:
            values_by_sub[row.sub_golongan] += _float_value(row.kwh)

    per_golongan = {key: 0 for key in KWH_JUAL_GROUP_LABELS}
    per_tegangan = {'TR': 0, 'TM': 0, 'TT': 0}
    detail = []
    for item in KWH_JUAL_CATALOG:
        kwh = values_by_sub.get(item['sub_golongan'], 0)
        per_golongan[item['group']] += kwh
        per_tegangan[item['tegangan']] += kwh
        detail.append({
            'group': item['group'],
            'group_label': KWH_JUAL_GROUP_LABELS[item['group']],
            'golongan': item['golongan'],
            'sub_golongan': item['sub_golongan'],
            'tegangan': item['tegangan'],
            'kwh': round(kwh, 3),
        })

    total = sum(per_tegangan.values())
    return {
        'periode': period.strftime('%Y-%m'),
        'periode_bulan': period.strftime('%Y-%m-%d'),
        'gi_id': gi_id,
        'catalog': _kwh_jual_catalog_payload(),
        'detail': detail,
        'per_golongan': {key: round(value, 3) for key, value in per_golongan.items()},
        'per_tegangan': {key: round(value, 3) for key, value in per_tegangan.items()},
        'total': round(total, 3),
        'trend': _kwh_jual_trend(gi_id, period),
    }


def _kwh_jual_trend(gi_id, period):
    start = _shift_month(period, -5)
    end = _next_month(period)
    q = KwhJual.query.filter(
        KwhJual.periode_bulan >= start,
        KwhJual.periode_bulan < end,
    )
    if gi_id:
        q = q.filter(KwhJual.gi_id == gi_id)
    monthly = {
        _shift_month(start, index).strftime('%Y-%m'): {'total': 0, 'TR': 0, 'TM': 0, 'TT': 0}
        for index in range(6)
    }
    for row in q.all():
        key = row.periode_bulan.strftime('%Y-%m')
        if key not in monthly:
            continue
        value = _float_value(row.kwh)
        monthly[key]['total'] += value
        if row.tegangan in {'TR', 'TM', 'TT'}:
            monthly[key][row.tegangan] += value
    return [
        {'periode': key, **{name: round(value, 3) for name, value in values.items()}}
        for key, values in monthly.items()
    ]


@app.route('/api/kwh-jual', methods=['GET', 'POST'])
def api_kwh_jual():
    try:
        if request.method == 'GET':
            gi_id = request.args.get('gi_id', type=int)
            bulan = (request.args.get('bulan') or request.args.get('periode') or '').strip()
            period = _month_date(bulan or date.today().strftime('%Y-%m'))
            return jsonify(_kwh_jual_payload(gi_id, period))

        denied = _master_writer_required()
        if denied:
            return denied
        payload = _request_payload()
        gi_id = int(payload.get('gi_id') or 0)
        gi = db.session.get(GarduInduk, gi_id)
        if not gi:
            return _json_error('Gardu induk wajib dipilih.', 400)
        period = _month_date(payload.get('bulan') or payload.get('periode') or payload.get('periode_bulan'))
        entries = payload.get('entries') or payload.get('detail') or []
        if not isinstance(entries, list):
            return _json_error('Format entries tidak valid.', 400)

        saved = 0
        total = Decimal('0')
        for item in entries:
            sub = _clean_value(item.get('sub_golongan'))
            catalog = KWH_JUAL_SUB_INDEX.get(sub)
            if not catalog:
                return _json_error(f'Sub-golongan tidak dikenali: {sub}', 400)
            kwh = Decimal(str(item.get('kwh') or 0))
            if kwh < 0:
                return _json_error(f'Nilai kWh tidak boleh negatif: {sub}', 400)
            row = KwhJual.query.filter_by(
                gi_id=gi.id,
                periode_bulan=period,
                sub_golongan=sub,
            ).first()
            if not row:
                row = KwhJual(
                    gi_id=gi.id,
                    periode_bulan=period,
                    sub_golongan=sub,
                )
                db.session.add(row)
            row.golongan = catalog['golongan']
            row.tegangan = catalog['tegangan']
            row.kwh = kwh
            saved += 1
            total += kwh

        _audit('UPSERT_KWH_JUAL', entity_type='kwh_jual', entity_id=f'{gi.id}:{period:%Y-%m}', detail={
            'gi_id': gi.id,
            'kode_gi': gi.kode_gi,
            'periode_bulan': period.strftime('%Y-%m-%d'),
            'rows': saved,
            'total_kwh': float(total),
        })
        db.session.commit()
        return jsonify(_kwh_jual_payload(gi.id, period))
    except ValueError as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@app.route('/api/upload', methods=['POST'])
@role_required('admin', 'operator')
def api_upload():
    """
    Upload file NKWh Excel/CSV.
    Implementasi lengkap ada di upload_engine.py.
    """
    if 'file' not in request.files:
        return jsonify({'error': 'Tidak ada file yang dikirim'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Nama file kosong'}), 400
    try:
        _check_upload_rate()
        _validate_upload_file(file, ALLOWED_GENERIC_UPLOADS)
    except ValueError as e:
        return jsonify({'error': str(e)}), 400

    # TODO: integrasikan dengan UploadEngine dari upload_engine.py
    # from upload_engine import UploadEngine
    # engine = UploadEngine()
    # result = engine.proses(tmp_path, gi_id, trafo_id, periode)
    return jsonify({'message': 'Upload endpoint aktif. Integrasi UploadEngine belum selesai.'}), 200


def _norm_col(value):
    return ''.join(ch for ch in str(value).strip().lower() if ch.isalnum())


def _pick(row, aliases, default=None):
    for alias in aliases:
        key = _norm_col(alias)
        if key in row and pd.notna(row[key]) and str(row[key]).strip() != '':
            return row[key]
    return default


def _num(value, default=0):
    if value is None or pd.isna(value) or str(value).strip() == '':
        return default
    text_value = str(value).replace('.', '').replace(',', '.') if isinstance(value, str) else value
    try:
        return float(text_value)
    except (TypeError, ValueError):
        return default


def _str_value(value, default=''):
    if value is None or pd.isna(value):
        return default
    return str(value).strip()


def _month_date(value, fallback=None):
    if value is None or pd.isna(value) or str(value).strip() == '':
        if fallback:
            year, month = fallback.split('-')[:2]
            return date(int(year), int(month), 1)
        raise ValueError('Kolom bulan/periode wajib diisi')
    if hasattr(value, 'year') and hasattr(value, 'month'):
        return date(int(value.year), int(value.month), 1)
    raw = str(value).strip()
    if len(raw) == 7 and raw[4] == '-':
        return date(int(raw[:4]), int(raw[5:7]), 1)
    parsed = pd.to_datetime(raw, errors='coerce', dayfirst=True)
    if pd.isna(parsed):
        raise ValueError(f'Format bulan tidak dikenali: {raw}')
    return date(int(parsed.year), int(parsed.month), 1)


def _previous_month(period):
    return date(period.year - 1, 12, 1) if period.month == 1 else date(period.year, period.month - 1, 1)


def _read_upload_table(file):
    _validate_upload_file(file, ALLOWED_GENERIC_UPLOADS)
    filename = file.filename.lower()
    if filename.endswith(('.xlsx', '.xlsm', '.xls')):
        frame = pd.read_excel(file)
    elif filename.endswith('.csv'):
        frame = pd.read_csv(file)
    else:
        raise ValueError('Format file harus CSV atau Excel (.xlsx/.xls)')
    frame = frame.dropna(how='all')
    if len(frame) > app.config['MAX_IMPORT_ROWS']:
        raise ValueError(f'Jumlah baris melebihi batas {app.config["MAX_IMPORT_ROWS"]}.')
    frame.columns = [_norm_col(col) for col in frame.columns]
    return frame


def _find_or_create_gi(row, default_gi_id):
    if default_gi_id:
        gi = GarduInduk.query.get(default_gi_id)
        if gi:
            return gi
    kode = _str_value(_pick(row, ['kode_gi', 'kode gardu induk', 'gi']), '')
    nama = _str_value(_pick(row, ['nama_gi', 'gardu_induk', 'gardu induk', 'nama gardu induk']), kode)
    if not kode and not nama:
        raise ValueError('GI tidak ditemukan. Isi kode_gi/nama_gi atau pilih default GI.')
    gi = GarduInduk.query.filter(
        (GarduInduk.kode_gi == kode) | (GarduInduk.nama_gi == nama)
    ).first()
    if gi:
        return gi
    gi = GarduInduk(kode_gi=kode or nama[:20].upper(), nama_gi=nama or kode, aktif=True)
    db.session.add(gi)
    db.session.flush()
    return gi


def _find_or_create_trafo(row, gi, default_trafo_id):
    if default_trafo_id:
        trafo = Trafo.query.get(default_trafo_id)
        if trafo:
            return trafo
    kode = _str_value(_pick(row, ['kode_trafo', 'trafo', 'kode trafo']), '')
    nama = _str_value(_pick(row, ['nama_trafo', 'nama trafo']), kode or 'Trafo 1')
    if not kode:
        kode = 'TRF-1'
    trafo = Trafo.query.filter_by(gi_id=gi.id, kode_trafo=kode).first()
    if trafo:
        return trafo
    trafo = Trafo(
        gi_id=gi.id,
        kode_trafo=kode,
        nama_trafo=nama or kode,
        kapasitas_mva=Decimal('0'),
        tegangan_kv=Decimal('20'),
        aktif=True,
    )
    db.session.add(trafo)
    db.session.flush()
    return trafo


def _find_or_create_penyulang(row, gi, trafo):
    kode = _str_value(_pick(row, ['kode_penyulang', 'kode penyulang', 'kode', 'penyulang']), '')
    nama = _str_value(_pick(row, ['nama_penyulang', 'nama penyulang', 'nama', 'feeder']), kode)
    if not kode:
        raise ValueError('kode_penyulang wajib diisi')
    penyulang = Penyulang.query.filter_by(trafo_id=trafo.id, kode_penyulang=kode).first()
    if not penyulang:
        penyulang = Penyulang(
            trafo_id=trafo.id,
            gi_id=gi.id,
            kode_penyulang=kode,
            nama_penyulang=nama or kode,
            aktif=True,
        )
        db.session.add(penyulang)
    penyulang.nama_penyulang = nama or penyulang.nama_penyulang
    penyulang.jenis = _str_value(_pick(row, ['jenis', 'jenis_penyulang'], penyulang.jenis or 'REGULAR'), 'REGULAR').upper()
    penyulang.area_up3 = _str_value(_pick(row, ['area_up3', 'area', 'up3', 'area up3'], penyulang.area_up3), penyulang.area_up3)
    penyulang.ex_cabang = _str_value(_pick(row, ['ex_cabang', 'ex cabang', 'cabang'], penyulang.ex_cabang), penyulang.ex_cabang)
    penyulang.status = _str_value(_pick(row, ['status', 'status_penyulang', 'status penyulang'], penyulang.status or 'AKTIF'), 'AKTIF').upper()
    penyulang.aktif = penyulang.status not in {'NONAKTIF', 'OFF', 'PADAM PERMANEN'}
    db.session.flush()
    return penyulang


def _reading_values(row):
    faktor = _num(_pick(row, ['faktor_kali', 'faktor', 'fk']), 1) or 1
    stand_awal = _num(_pick(row, ['stand_awal', 'stand awal', 'awal']), 0)
    stand_akhir = _num(_pick(row, ['stand_akhir', 'stand akhir', 'akhir']), 0)
    wbp = _num(_pick(row, ['kwh_wbp', 'wbp']), 0)
    lwbp1 = _num(_pick(row, ['kwh_lwbp1', 'lwbp1', 'lwbp', 'lwbp_1']), 0)
    lwbp2 = _num(_pick(row, ['kwh_lwbp2', 'lwbp2', 'lwbp_2']), 0)
    total = _num(_pick(row, ['kwh_total', 'total_kwh', 'total kwh', 'total']), 0)

    if stand_akhir and faktor and not any([wbp, lwbp1, lwbp2, total]):
        total = max(0, (stand_akhir - stand_awal) * faktor)
        wbp, lwbp1, lwbp2 = total, 0, 0
    elif total and not any([wbp, lwbp1, lwbp2]):
        wbp, lwbp1, lwbp2 = total, 0, 0
    else:
        total = wbp + lwbp1 + lwbp2

    return stand_awal, stand_akhir, faktor, wbp, lwbp1, lwbp2, total


def _set_anomaly(reading, threshold_pct, min_delta):
    previous = FeederReading.query.filter_by(
        penyulang_id=reading.penyulang_id,
        periode_bulan=_previous_month(reading.periode_bulan)
    ).first()
    reading.flag_alert = False
    reading.deviasi_persen = Decimal('0')
    reading.anomaly_type = None
    reading.catatan = None
    if not previous or previous.kwh_total <= 0:
        return
    delta = reading.kwh_total - previous.kwh_total
    pct = (delta / previous.kwh_total) * 100
    if abs(delta) >= min_delta and abs(pct) >= threshold_pct:
        reading.flag_alert = True
        reading.deviasi_persen = Decimal(str(round(pct, 2)))
        reading.anomaly_type = 'NAIK' if delta > 0 else 'TURUN'
        reading.catatan = (
            f'Anomali {reading.anomaly_type.lower()} {round(pct, 2)}% '
            f'dari bulan sebelumnya ({round(previous.kwh_total)} ke {round(reading.kwh_total)} kWh).'
        )


def _slug_code(value, fallback='DATA', max_len=30):
    raw = _str_value(value, fallback).upper()
    code = ''.join(ch if ch.isalnum() else '-' for ch in raw)
    code = '-'.join(part for part in code.split('-') if part)
    return (code or fallback)[:max_len]


def _decimal_or_none(value):
    if value is None:
        return None
    return Decimal(str(value))


def _decimal_or_zero(value):
    return Decimal(str(value or 0))


def _nkwh_period(value, fallback=None):
    if value:
        return date.fromisoformat(value)
    if fallback:
        return _month_date(fallback)
    raise ValueError('Periode bulan tidak ditemukan dari workbook. Isi default bulan saat import.')


def _find_or_create_gi_from_name(name):
    nama = _str_value(name, 'Belum Dipetakan')
    kode = _slug_code(nama, 'GI', 20)
    gi = GarduInduk.query.filter(
        (GarduInduk.kode_gi == kode) | (GarduInduk.nama_gi == nama)
    ).first()
    if gi:
        return gi
    gi = GarduInduk(kode_gi=kode, nama_gi=nama, aktif=True)
    db.session.add(gi)
    db.session.flush()
    return gi


def _find_or_create_trafo_from_nkwh(gi, kode_trafo, nama_trafo):
    raw = _str_value(kode_trafo, 'TRF-1')
    raw_upper = raw.upper()
    kode = raw_upper if raw_upper.startswith('TRF') else f'TRF-{raw_upper}'
    candidates = {raw_upper, kode, f'{gi.kode_gi}-T{raw_upper}', f'{gi.kode_gi}-{kode}'}
    trafo = Trafo.query.filter(Trafo.gi_id == gi.id, Trafo.kode_trafo.in_(candidates)).first()
    if not trafo:
        trafo = Trafo.query.filter(
            Trafo.gi_id == gi.id,
            Trafo.nama_trafo.in_({_str_value(nama_trafo, kode), f'Trafo {raw}'})
        ).first()
    if trafo:
        return trafo
    trafo = Trafo(
        gi_id=gi.id,
        kode_trafo=kode,
        nama_trafo=_str_value(nama_trafo, kode),
        kapasitas_mva=Decimal('0'),
        tegangan_kv=Decimal('20'),
        aktif=True,
    )
    db.session.add(trafo)
    db.session.flush()
    return trafo


def _find_or_create_penyulang_from_nkwh(item, gi, trafo):
    kode = _slug_code(item.get('kode_penyulang') or item.get('nama_penyulang'), 'PENYULANG', 30)
    nama = _str_value(item.get('nama_penyulang'), kode)
    penyulang = Penyulang.query.filter_by(trafo_id=trafo.id, kode_penyulang=kode).first()
    if not penyulang:
        penyulang = Penyulang.query.filter_by(trafo_id=trafo.id, nama_penyulang=nama).first()
    if not penyulang:
        penyulang = Penyulang(
            trafo_id=trafo.id,
            gi_id=gi.id,
            kode_penyulang=kode,
            nama_penyulang=nama,
            jenis='REGULAR',
            status='AKTIF',
            aktif=True,
        )
        db.session.add(penyulang)
    penyulang.gi_id = gi.id
    penyulang.nama_penyulang = nama or penyulang.nama_penyulang
    penyulang.status = penyulang.status or 'AKTIF'
    penyulang.aktif = penyulang.status not in {'NONAKTIF', 'OFF', 'PADAM PERMANEN'}
    db.session.flush()
    return penyulang


def _apply_nkwh_registers(reading, item):
    registers = item.get('registers') or {}
    for prefix in ('wbp', 'lwbp1', 'lwbp2'):
        detail = registers.get(prefix, {})
        setattr(reading, f'{prefix}_stand_awal', _decimal_or_none(detail.get('stand_awal')))
        setattr(reading, f'{prefix}_stand_akhir', _decimal_or_none(detail.get('stand_akhir')))
        setattr(reading, f'{prefix}_faktor_kali', _decimal_or_none(detail.get('faktor_kali')))

    first_register = next((registers[key] for key in ('wbp', 'lwbp1', 'lwbp2') if key in registers), {})
    reading.stand_awal = _decimal_or_none(first_register.get('stand_awal'))
    reading.stand_akhir = _decimal_or_none(first_register.get('stand_akhir'))
    reading.faktor_kali = _decimal_or_none(first_register.get('faktor_kali')) or Decimal('1')

    reading.kwh_wbp = _decimal_or_zero(item.get('kwh_wbp'))
    reading.kwh_lwbp1 = _decimal_or_zero(item.get('kwh_lwbp1'))
    reading.kwh_lwbp2 = _decimal_or_zero(item.get('kwh_lwbp2'))
    reading.manual_kwh_wbp = _decimal_or_none(item.get('manual_kwh_wbp'))
    reading.manual_kwh_lwbp1 = _decimal_or_none(item.get('manual_kwh_lwbp1'))
    reading.manual_kwh_lwbp2 = _decimal_or_none(item.get('manual_kwh_lwbp2'))
    reading.source_format = 'NKWH_XLSX'
    reading.source_sheet = item.get('source_sheet')
    reading.source_row_start = item.get('source_row_start')
    reading.source_row_end = item.get('source_row_end')


def _import_nkwh_exim_rows(exim_rows, period):
    imported = updated = 0
    for row in exim_rows:
        kode_rule = _slug_code(
            f"{row.get('gardu_induk')}-{row.get('feeder')}-{row.get('row')}",
            'EXIM',
            60
        )
        rule = EximRule.query.filter_by(kode_rule=kode_rule).first()
        if rule:
            updated += 1
        else:
            imported += 1
            rule = EximRule(kode_rule=kode_rule)
            db.session.add(rule)

        rule.nama_rule = _str_value(row.get('feeder'), kode_rule)
        rule.metode = row.get('metode') or 'ADJUSTMENT'
        rule.up3_asal = row.get('area_asal') or None
        rule.up3_tujuan = row.get('area_tujuan') or None
        rule.fungsi = row.get('fungsi') or None
        rule.arah = row.get('arah') or None
        rule.periode_mulai = period
        rule.source_sheet = 'Exim'
        rule.source_row = row.get('row')
        rule.catatan = row.get('jenis') or row.get('lokasi')
        db.session.flush()

        result = EximMonthlyResult.query.filter_by(
            rule_id=rule.id,
            periode_bulan=period,
            up3_tujuan=rule.up3_tujuan,
        ).first()
        if not result:
            result = EximMonthlyResult(
                rule_id=rule.id,
                periode_bulan=period,
            )
            db.session.add(result)

        basis = row.get('kwh_penyulang_basis') or 0
        transfer = row.get('kwh_total') or sum([
            row.get('kwh_wbp') or 0,
            row.get('kwh_lwbp1') or 0,
            row.get('kwh_lwbp2') or 0,
        ])
        result.metode = rule.metode
        result.up3_asal = rule.up3_asal
        result.up3_tujuan = rule.up3_tujuan
        result.fungsi = rule.fungsi
        result.arah = rule.arah
        result.kwh_basis = _decimal_or_zero(basis)
        result.kwh_wbp = _decimal_or_zero(row.get('kwh_wbp'))
        result.kwh_lwbp1 = _decimal_or_zero(row.get('kwh_lwbp1'))
        result.kwh_lwbp2 = _decimal_or_zero(row.get('kwh_lwbp2'))
        result.kwh_transfer = _decimal_or_zero(transfer)
        result.porsi = _decimal_or_none(transfer / basis) if basis else None
        result.source_sheet = 'Exim'
        result.source_row = row.get('row')
        result.catatan = row.get('lokasi')
    return imported, updated


def _nkwh_import_blockers(parsed):
    blockers = []
    if not parsed.get('feeder_count'):
        blockers.append('Tidak ada data penyulang yang bisa diimport.')
    return blockers


@app.route('/api/nkwh/analyze', methods=['POST'])
@role_required('admin', 'operator')
def api_nkwh_analyze():
    if 'file' not in request.files:
        return jsonify({'error': 'Tidak ada file yang dikirim'}), 400
    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'Nama file kosong'}), 400

    try:
        _check_upload_rate()
        safe_filename, _ = _validate_upload_file(file, ALLOWED_NKWH_UPLOADS)
        result = analyze_workbook(file.stream)
        if result.get('kwh_penyulang', {}).get('feeder_count', 0) > app.config['MAX_IMPORT_ROWS']:
            return jsonify({'error': f'Jumlah data penyulang melebihi batas {app.config["MAX_IMPORT_ROWS"]}.'}), 400
        result['filename'] = safe_filename
        default_bulan = request.form.get('bulan', '').strip()
        period_value = result.get('periode_bulan') or default_bulan
        if period_value:
            period = _month_date(period_value)
            result['workflow'] = _workflow_payload(period)
        _audit('ANALYZE_NKWH', entity_type='upload', detail={
            'filename': safe_filename,
            'periode_bulan': result.get('periode_bulan'),
            'feeder_count': result.get('kwh_penyulang', {}).get('feeder_count'),
            'exim_rows': result.get('exim', {}).get('row_count'),
        })
        db.session.commit()
        return jsonify(result)
    except ValueError as e:
        db.session.rollback()
        _safe_commit_audit('ANALYZE_NKWH', detail={'filename': file.filename, 'error': str(e)}, status='FAILED')
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        db.session.rollback()
        _safe_commit_audit('ANALYZE_NKWH', detail={'filename': file.filename, 'error': str(e)}, status='FAILED')
        return jsonify({'error': str(e)}), 500


@app.route('/api/nkwh/import', methods=['POST'])
@role_required('admin', 'operator')
def api_nkwh_import():
    if 'file' not in request.files:
        return jsonify({'error': 'Tidak ada file yang dikirim'}), 400
    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'Nama file kosong'}), 400

    threshold_pct = request.form.get('threshold_pct', default=25, type=float)
    min_delta = request.form.get('min_delta', default=10000, type=float)
    default_bulan = request.form.get('bulan', '').strip()
    import_exim = request.form.get('import_exim', '1') == '1'

    try:
        _check_upload_rate()
        safe_filename, _ = _validate_upload_file(file, ALLOWED_NKWH_UPLOADS)
        parsed = parse_nkwh_feeders(file.stream)
        if parsed.get('feeder_count', 0) > app.config['MAX_IMPORT_ROWS']:
            return jsonify({'error': f'Jumlah data penyulang melebihi batas {app.config["MAX_IMPORT_ROWS"]}.'}), 400
        blockers = _nkwh_import_blockers(parsed)
        if blockers:
            return jsonify({'error': 'Import dibatalkan karena validasi gagal.', 'errors': blockers}), 400
        period = _nkwh_period(parsed.get('periode_bulan'), default_bulan or None)
        _ensure_period_writable(period)
        created = updated = alerts = 0

        for item in parsed.get('feeders', []):
            gi = _find_or_create_gi_from_name(item.get('gardu_induk'))
            trafo = _find_or_create_trafo_from_nkwh(gi, item.get('kode_trafo'), item.get('nama_trafo'))
            penyulang = _find_or_create_penyulang_from_nkwh(item, gi, trafo)

            reading = FeederReading.query.filter_by(
                penyulang_id=penyulang.id,
                periode_bulan=period,
            ).first()
            if reading:
                updated += 1
            else:
                created += 1
                reading = FeederReading(
                    penyulang_id=penyulang.id,
                    periode_bulan=period,
                )
                db.session.add(reading)

            reading.trafo_id = trafo.id
            reading.gi_id = gi.id
            _apply_nkwh_registers(reading, item)
            db.session.flush()
            _set_anomaly(reading, threshold_pct, min_delta)
            if reading.flag_alert:
                alerts += 1

        exim_created = exim_updated = 0
        if import_exim:
            file.stream.seek(0)
            exim = parse_exim_rows(file.stream)
            exim_created, exim_updated = _import_nkwh_exim_rows(exim.get('rows', []), period)

        workflow_record = _mark_period_uploaded(period, 'NKWH', safe_filename)
        _audit('IMPORT_NKWH', entity_type='upload', detail={
            'filename': safe_filename,
            'periode_bulan': period.strftime('%Y-%m-%d'),
            'created': created,
            'updated': updated,
            'alerts': alerts,
            'exim_created': exim_created,
            'exim_updated': exim_updated,
            'feeder_count': parsed.get('feeder_count', 0),
            'gi_count': parsed.get('gi_count', 0),
            'workflow_status': workflow_record.status,
        })
        db.session.commit()
        return jsonify({
            'message': 'Import NKWh selesai',
            'periode_bulan': period.strftime('%Y-%m-%d'),
            'created': created,
            'updated': updated,
            'alerts': alerts,
            'exim_created': exim_created,
            'exim_updated': exim_updated,
            'feeder_count': parsed.get('feeder_count', 0),
            'gi_count': parsed.get('gi_count', 0),
            'total_kwh': parsed.get('total_kwh', 0),
            'workflow': _workflow_payload(period, workflow_record),
        })
    except ValueError as e:
        db.session.rollback()
        _safe_commit_audit('IMPORT_NKWH', detail={'filename': file.filename, 'error': str(e)}, status='FAILED')
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        db.session.rollback()
        _safe_commit_audit('IMPORT_NKWH', detail={'filename': file.filename, 'error': str(e)}, status='FAILED')
        return jsonify({'error': str(e)}), 500


@app.route('/api/upload-penyulang', methods=['POST'])
@role_required('admin', 'operator')
def api_upload_penyulang():
    if 'file' not in request.files:
        return jsonify({'error': 'Tidak ada file yang dikirim'}), 400
    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'Nama file kosong'}), 400

    default_gi_id = request.form.get('gi_id', type=int)
    default_trafo_id = request.form.get('trafo_id', type=int)
    default_bulan = request.form.get('bulan', '').strip()
    threshold_pct = request.form.get('threshold_pct', default=25, type=float)
    min_delta = request.form.get('min_delta', default=10000, type=float)

    try:
        _check_upload_rate()
        safe_filename, _ = _validate_upload_file(file, ALLOWED_GENERIC_UPLOADS)
        frame = _read_upload_table(file)
        created = updated = alerts = 0
        errors = []
        checked_periods = set()
        imported_periods = {}

        for idx, raw in frame.iterrows():
            row = raw.to_dict()
            try:
                period = _month_date(_pick(row, ['bulan', 'periode', 'periode_bulan', 'month']), default_bulan or None)
                period_key = period.isoformat()
                if period_key not in checked_periods:
                    _ensure_period_writable(period)
                    checked_periods.add(period_key)
                gi = _find_or_create_gi(row, default_gi_id)
                trafo = _find_or_create_trafo(row, gi, default_trafo_id)
                penyulang = _find_or_create_penyulang(row, gi, trafo)
                stand_awal, stand_akhir, faktor, wbp, lwbp1, lwbp2, total = _reading_values(row)

                reading = FeederReading.query.filter_by(
                    penyulang_id=penyulang.id,
                    periode_bulan=period
                ).first()
                if reading:
                    updated += 1
                else:
                    created += 1
                    reading = FeederReading(
                        penyulang_id=penyulang.id,
                        trafo_id=trafo.id,
                        gi_id=gi.id,
                        periode_bulan=period,
                    )
                    db.session.add(reading)

                reading.trafo_id = trafo.id
                reading.gi_id = gi.id
                reading.stand_awal = Decimal(str(stand_awal))
                reading.stand_akhir = Decimal(str(stand_akhir))
                reading.faktor_kali = Decimal(str(faktor))
                reading.kwh_wbp = Decimal(str(wbp))
                reading.kwh_lwbp1 = Decimal(str(lwbp1))
                reading.kwh_lwbp2 = Decimal(str(lwbp2))
                db.session.flush()
                _set_anomaly(reading, threshold_pct, min_delta)
                if reading.flag_alert:
                    alerts += 1
                imported_periods[period_key] = period
            except Exception as row_error:
                errors.append({'baris': int(idx) + 2, 'error': str(row_error)})

        if errors and not (created or updated):
            db.session.rollback()
            _safe_commit_audit('IMPORT_PENYULANG', detail={
                'filename': safe_filename,
                'errors': errors[:10],
            }, status='FAILED')
            return jsonify({'error': 'Upload gagal. Tidak ada baris valid.', 'errors': errors[:10]}), 400

        workflow_rows = []
        for period in imported_periods.values():
            record = _mark_period_uploaded(period, 'PENYULANG', safe_filename)
            workflow_rows.append(_workflow_payload(period, record))

        _audit('IMPORT_PENYULANG', entity_type='upload', detail={
            'filename': safe_filename,
            'created': created,
            'updated': updated,
            'alerts': alerts,
            'error_count': len(errors),
            'periods': sorted(imported_periods),
        })
        db.session.commit()
        return jsonify({
            'message': 'Upload penyulang selesai',
            'created': created,
            'updated': updated,
            'alerts': alerts,
            'errors': errors[:10],
            'error_count': len(errors),
            'workflow': workflow_rows,
        })
    except ValueError as e:
        db.session.rollback()
        _safe_commit_audit('IMPORT_PENYULANG', detail={'filename': file.filename, 'error': str(e)}, status='FAILED')
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        db.session.rollback()
        _safe_commit_audit('IMPORT_PENYULANG', detail={'filename': file.filename, 'error': str(e)}, status='FAILED')
        return jsonify({'error': str(e)}), 500


# ════════════════════════════════════════════════
# ERROR HANDLERS
# ════════════════════════════════════════════════

@app.errorhandler(403)
def forbidden(e):
    if _wants_json():
        return jsonify({'error': 'Akses ditolak.'}), 403
    return render_template('error.html',
                           kode=403,
                           judul='Akses Ditolak',
                           pesan='Kamu tidak memiliki izin untuk membuka halaman ini.'), 403


@app.errorhandler(413)
def payload_too_large(e):
    max_mb = app.config['MAX_CONTENT_LENGTH'] // (1024 * 1024)
    if _wants_json():
        return jsonify({'error': f'Ukuran file melebihi batas {max_mb} MB.'}), 413
    return render_template('error.html',
                           kode=413,
                           judul='File Terlalu Besar',
                           pesan=f'Ukuran file melebihi batas {max_mb} MB.'), 413


@app.errorhandler(404)
def not_found(e):
    return render_template('error.html',
                           kode=404,
                           judul='Halaman Tidak Ditemukan',
                           pesan='URL yang kamu akses tidak ada.'), 404

@app.errorhandler(500)
def server_error(e):
    return render_template('error.html',
                           kode=500,
                           judul='Server Error',
                           pesan='Terjadi kesalahan di server. Cek log untuk detail.'), 500


# ════════════════════════════════════════════════
# JALANKAN
# ════════════════════════════════════════════════

if __name__ == '__main__':
    debug_mode = os.getenv('FLASK_DEBUG', 'false').lower() == 'true'
    app.run(debug=debug_mode, host='0.0.0.0', port=5000)

import logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
